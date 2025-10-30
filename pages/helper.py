import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import zipfile
import os
import csv
import xlrd
import openpyxl
import os
import pandas as pd


def output_save_in_template(data, headers, filename="Report.xlsx"):
    """
    Create an Excel workbook in memory with styled headers and zebra-striping.
    
    Args:
        data (list[dict]): List of row dicts
        headers (list[str]): Header column names
        filename (str): Desired filename inside the zip (default "Report.xlsx")
    
    Returns:
        (BytesIO, str): Excel file content in memory and its filename
    """
    # Create workbook directly (faster than loading from template)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Define styles once (optimization)
    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    zebra_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    # Write headers quickly
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font

    # Pre-build header map for faster lookups
    header_map = {header: idx + 1 for idx, header in enumerate(headers)}
    start_row = 2

    # Optimized data writing with fast zebra striping
    # First pass: Write all data quickly
    for i, record in enumerate(data, start=start_row):
        for header in headers:
            col = header_map[header]
            ws.cell(row=i, column=col, value=record.get(header, ""))
    
    # Second pass: Apply zebra striping in batches (much faster)
    if data:  # Only if we have data
        max_row = start_row + len(data) - 1
        max_col = len(headers)
        
        # Apply zebra striping to even rows only (starting from row 2)
        for row in range(start_row, max_row + 1, 2):  # Every 2nd row starting from start_row
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = zebra_fill

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename

def multiple_excels_to_zip(excel_files, zip_filename="Output.zip"):
    """
    Create a ZIP archive from multiple Excel files (all in memory) with optimized compression.
    
    Args:
        excel_files (list of tuples): [(excel_io, excel_filename), ...]
            - excel_io: BytesIO containing Excel content
            - excel_filename: Name for the file inside the zip
    
    Returns:
        BytesIO: In-memory ZIP archive
    """
    zip_buffer = io.BytesIO()
    # Use faster compression level (1-3 instead of default 6)
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED, compresslevel=1) as zf:
        for excel_io, excel_filename in excel_files:
            excel_io.seek(0)  # make sure pointer is at start
            zf.writestr(excel_filename, excel_io.read())

    zip_buffer.seek(0)
    return zip_buffer


def read_file(
    file_path: str,
    sheet_name: str | None = None,
    start_row: int = 0,
    header: bool = True,
    skip_blank_rows: bool = True,
    **kwargs
) -> pd.DataFrame:
    """
    Dynamic pandas read function for CSV/XLS/XLSX with flexible options.

    Args:
        file_path (str): Path of the file
        sheet_name (str | None): Sheet name (for Excel files)
        start_row (int): Row index to start reading (0-based)
        header (bool): Whether to use the first row after start_row as column names
        skip_blank_rows (bool): Whether to skip completely empty rows
        kwargs: Additional pandas read_* options (usecols, dtype, parse_dates, etc.)

    Returns:
        pd.DataFrame
    """
    ext = os.path.splitext(file_path)[-1].lower()
    
    # Determine header row
    if header:
        # Use the row as header
        header_row = start_row
        skiprows = None
    else:
        # Treat all rows as data (assign default column names or user-provided)
        header_row = None
        skiprows = start_row

    if ext in [".xlsx", ".xls"]:
        df = pd.read_excel(file_path, sheet_name=sheet_name,
                           header=header_row, skiprows=skiprows, **kwargs)
        
        # Handle case where read_excel returns a dict (multiple sheets)
        if isinstance(df, dict):
            # Get the first sheet if no specific sheet was requested
            df = list(df.values())[0]
            
    elif ext == ".csv":
        df = pd.read_csv(file_path,
                         header=header_row, skiprows=skiprows, **kwargs)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

    # Remove blank rows if requested
    if skip_blank_rows:
        # Drop rows where all values are NaN/None/empty
        df = df.dropna(how='all')
        # Reset index after dropping rows
        df = df.reset_index(drop=True)

    return df