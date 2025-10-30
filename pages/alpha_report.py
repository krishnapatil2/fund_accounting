import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import json
import zipfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border
from decimal import Decimal
from CONSTANTS import *
import pandas as pd
import threading


from my_app.pages.loading import LoadingSpinner
from .helper import output_save_in_template, multiple_excels_to_zip


def _safe_decimal(value):
    if value in [None, ""]:
        return Decimal("0")
    cleaned = str(value).replace(",", "").strip()
    try:
        return Decimal(cleaned)
    except Exception:
        return Decimal("0")


def _format_date(date_str: str):
    # '18Sep25' -> (yyyymmdd, mm-dd-YYYY)
    date_obj = datetime.strptime(date_str, '%d%b%y')
    yyyymmdd = date_obj.strftime('%Y%m%d')
    mm_dd_yyyy = date_obj.strftime('%m-%d-%Y')
    return yyyymmdd, mm_dd_yyyy


class AlphaReportPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg="#ecf0f1")

        # Title
        title = tk.Label(self, text="📑 Alpha Report", font=("Arial", 20, "bold"), bg="#ecf0f1", fg="#2c3e50")
        title.pack(pady=10)

        # Controls
        controls = tk.Frame(self, bg="#ecf0f1")
        controls.pack(fill="x", padx=20, pady=5)

        self.csv_path_var = tk.StringVar()

        tk.Label(controls, text="Trade CSV:", font=("Arial", 11), bg="#ecf0f1", fg="#2c3e50").pack(side="left")
        tk.Entry(controls, textvariable=self.csv_path_var, width=60).pack(side="left", padx=8)
        tk.Button(controls, text="Browse", command=self._browse_csv, bg="#3498db", fg="white", relief="flat", padx=10, pady=4).pack(side="left")

        tk.Button(controls, text="Process", command=self._process, bg="#27ae60", fg="white", relief="flat", padx=14, pady=6, font=("Arial", 11, "bold")).pack(side="left", padx=10)
        tk.Button(controls, text="Export Excel", command=self._export_excel, bg="#8e44ad", fg="white", relief="flat", padx=14, pady=6, font=("Arial", 11, "bold")).pack(side="left")
        tk.Button(controls, text="Export to Template", command=self._export_to_template, bg="#d35400", fg="white", relief="flat", padx=14, pady=6, font=("Arial", 11, "bold")).pack(side="left", padx=10)

        # Status
        self.status_var = tk.StringVar(value="Load a CSV and click Process")
        tk.Label(self, textvariable=self.status_var, font=("Arial", 10), bg="#ecf0f1", fg="#7f8c8d").pack(fill="x", padx=20)

        # Split views
        content = tk.Frame(self, bg="#ecf0f1")
        content.pack(fill="both", expand=True, padx=20, pady=10)

        left = tk.Frame(content, bg="#ecf0f1")
        right = tk.Frame(content, bg="#ecf0f1")
        left.pack(side="left", fill="both", expand=True, padx=(0, 10))
        right.pack(side="left", fill="both", expand=True)

        # For_Trade search + table
        header_left = tk.Frame(left, bg="#ecf0f1")
        header_left.pack(fill="x", pady=(0, 4))
        tk.Label(header_left, text="For Trade", font=("Arial", 13, "bold"), bg="#ecf0f1", fg="#2c3e50").pack(side="left")
        # left search aligned to right of its header (label then entry)
        self.for_trade_search_var = tk.StringVar()
        self.for_trade_search_var.trace_add("write", lambda *_: self._render_for_trade())
        left_search_box = tk.Frame(header_left, bg="#ecf0f1")
        left_search_box.pack(side="right", padx=(0, 18))
        tk.Label(left_search_box, text="Search:", font=("Arial", 10), bg="#ecf0f1", fg="#2c3e50").pack(side="left")
        tk.Entry(left_search_box, textvariable=self.for_trade_search_var, width=24).pack(side="left", padx=(6, 0))
        left_table = tk.Frame(left, bg="#ecf0f1")
        left_table.pack(fill="both", expand=True)
        self.for_trade_tree = ttk.Treeview(left_table, columns=("Scrip", "Buy/Sell", "Qty", "Rate"), show="headings", height=12)
        for col, w in [("Scrip", 260), ("Buy/Sell", 90), ("Qty", 100), ("Rate", 100)]:
            self.for_trade_tree.heading(col, text=col)
            self.for_trade_tree.column(col, width=w, anchor="w")
        y1 = ttk.Scrollbar(left_table, orient="vertical", command=self.for_trade_tree.yview)
        x1 = ttk.Scrollbar(left_table, orient="horizontal", command=self.for_trade_tree.xview)
        self.for_trade_tree.configure(yscrollcommand=y1.set, xscrollcommand=x1.set)
        # grid layout so the bottom scrollbar spans full width
        self.for_trade_tree.grid(row=0, column=0, sticky="nsew")
        y1.grid(row=0, column=1, sticky="ns")
        x1.grid(row=1, column=0, columnspan=2, sticky="ew")
        left_table.rowconfigure(0, weight=1)
        left_table.columnconfigure(0, weight=1)

        # Security_Creation search + table
        header_right = tk.Frame(right, bg="#ecf0f1")
        header_right.pack(fill="x", pady=(0, 4))
        tk.Label(header_right, text="Security Creation", font=("Arial", 13, "bold"), bg="#ecf0f1", fg="#2c3e50").pack(side="left")
        self.sec_search_var = tk.StringVar()
        self.sec_search_var.trace_add("write", lambda *_: self._render_sec())
        right_search_box = tk.Frame(header_right, bg="#ecf0f1")
        right_search_box.pack(side="right", padx=(0, 18))
        tk.Label(right_search_box, text="Search:", font=("Arial", 10), bg="#ecf0f1", fg="#2c3e50").pack(side="left")
        tk.Entry(right_search_box, textvariable=self.sec_search_var, width=24).pack(side="left", padx=(6, 0))
        right_table = tk.Frame(right, bg="#ecf0f1")
        right_table.pack(fill="both", expand=True)
        self.sec_tree = ttk.Treeview(
            right_table,
            columns=("Scrip Name", "Underlying", "Expiry", "Strike Price", "CallPut", "Trading Size"),
            show="headings",
            height=12
        )
        for col, w in [
            ("Scrip Name", 260), ("Underlying", 100), ("Expiry", 140), ("Strike Price", 110), ("CallPut", 80), ("Trading Size", 110)
        ]:
            self.sec_tree.heading(col, text=col)
            self.sec_tree.column(col, width=w, anchor="w")
        y2 = ttk.Scrollbar(right_table, orient="vertical", command=self.sec_tree.yview)
        x2 = ttk.Scrollbar(right_table, orient="horizontal", command=self.sec_tree.xview)
        self.sec_tree.configure(yscrollcommand=y2.set, xscrollcommand=x2.set)
        # grid layout so the bottom scrollbar spans full width
        self.sec_tree.grid(row=0, column=0, sticky="nsew")
        y2.grid(row=0, column=1, sticky="ns")
        x2.grid(row=1, column=0, columnspan=2, sticky="ew")
        right_table.rowconfigure(0, weight=1)
        right_table.columnconfigure(0, weight=1)

        # Data holders
        self.for_trade_rows = []
        self.security_creation_rows = []
        self._aafspl_car_future_data = []
        self._car_trade_loader_data = []
        self._option_security_data = []

    # ---- UI handlers ----
    def _browse_csv(self):
        path = filedialog.askopenfilename(title="Select Trade CSV", filetypes=[["CSV Files", "*.csv"], ["All Files", "*.*"]])
        if path:
            self.csv_path_var.set(path)


    # ---- Core processing (reflecting test.py) ----
    def _process(self):
        # Clear tables
        for it in self.for_trade_tree.get_children():
            self.for_trade_tree.delete(it)
        for it in self.sec_tree.get_children():
            self.sec_tree.delete(it)
        self.for_trade_rows = []
        self.security_creation_rows = []
        self._aafspl_car_future_data = []
        self._car_trade_loader_data = []
        self._option_security_data = []

        # Load consolidated JSON
        from my_app.file_utils import get_app_directory
        app_dir = get_app_directory()
        consolidated_path = os.path.join(app_dir, "consolidated_data.json")

        try:
            if os.path.exists(consolidated_path):
                with open(consolidated_path, "r") as f:
                    consolidated_data = json.load(f)
            else:
                # Create default consolidated data if file doesn't exist
                consolidated_data = {
                    "lotsize_data": {
                        "ALUMINI": 1000,
                        "ALUMINIUM": 5000,
                        "CASTOR": 50,
                        "COCUDAKL": 100,
                        "COPPER": 2500, 
                        "CRUDEOIL": 100,
                        "CRUDEOILM": 10,
                        "DHANIYA": 50,
                        "GOLD": 100,
                        "GOLDM": 10,
                        "GUARGUM5": 50,
                        "GUARSEED10": 50,
                        "JEERAUNJHA": 30,
                        "LEAD": 5000,
                        "LEADMINI": 1000,
                        "NATGASMINI": 250,
                        "NATGAS": 1250,
                        "SILVER": 30,
                        "SILVERM": 5,
                        "SILVERMIC": 1,
                        "TMCFGRNZM": 50,
                        "ZINC": 5000,
                        "ZINCMINI": 1000,
                        "JEERAMINI": 10,
                        "KAPAS": 200,
                        "MCXBULLDEX": 30,
                        "WTICRUDE": 100,
                        "NATURALGAS": 1250,
                        "GOLDTEN": 1,
                        "ELECDMBL": 50,
                        "ELECMBL": 50
                    },
                    "trade_headers": {
                        "CLIENT_CODE": "Client\nCode",
                        "EXCHANGE": "Exchange",
                        "MKT_TYPE": "Mkt\nType",
                        "DATE": "Date",
                        "TRADE_TIME": "TradeTime",
                        "SETL_NO": "SetlNo",
                        "SCRIP_CODE": "Scrip Code",
                        "SCRIP_NAME": "Scrip Name",
                        "ISIN": "ISIN",
                        "BUY_QTY": "Buy Qty",
                        "BUY_RATE": "Buy Rate",
                        "BUY_NET_RATE": "Buy\nNet Rate",
                        "BUY_VALUE": "Buy Value",
                        "SELL_QTY": "Sell Qty",
                        "BALANCE": "Balance",
                        "SELL_RATE": "Sell Rate",
                        "SELL_NET_RATE": "Sell\nNet Rate",
                        "SELL_VALUE": "Sell Value",
                        "BREAK_EVEN": "BreakEven",
                        "MARKET_RATE": "Market Rate",
                        "TERMINAL": "Terminal",
                        "DELY": "Dely"
                    },
                    "aafspl_car_future": {
                        "Exchange": "MCX",
                        "Issuer": "",
                        "Ticker": "",
                        "Cusip": "",
                        "Sedol": "",
                        "Isin": "",
                        "AltKey1": "",
                        "AltKey2": "",
                        "BloombergID": "",
                        "BloombergTicker": "",
                        "BloombergUniqueID": "",
                        "BloombergMarketSector": "",
                        "SettleDays": "",
                        "PricingFactor": 1,
                        "CurrentPriceDayRange": 1,
                        "AssetType": "FT",
                        "InvestmentType": "CMFT",
                        "PriceDenomination": "INR",
                        "BifurcationCurrency": "INR",
                        "PrincipalCurrency": "INR",
                        "IncomeCurrency": "INR",
                        "RiskCurrency": "INR",
                        "IssueCountry": "IN",
                        "WithholdingTaxType": "Standard",
                        "QDIEligibilityFlag": "Not Eligible",
                        "SharesOutstanding": "",
                        "SubIndustry": "",
                        "SubIndustry2": "",
                        "QuantityPrecision": 3,
                        "InvestmentCrossZero": "Use Accounting Parameters",
                        "PriceByPreference": "Currency",
                        "ForwardPriceInterpolateFlag": 0,
                        "PricingPrecision": 3,
                        "FirstMarkDate": "01-01-2022",
                        "LastMarkDate": "",
                        "PriceList": "",
                        "AutoGenerateMarks": 1,
                        "CashSettlement": 1
                    },
                    "option_security": {
                        "Exchange": "MCX",
                        "Issuer": "",
                        "Ticker": "",
                        "Cusip": "",
                        "Sedol": "",
                        "Isin": "",
                        "Riccode": "",
                        "AltKey1": "",
                        "AltKey2": "",
                        "BloombergID": "",
                        "BloombergTicker": "",
                        "BloombergUniqueID": "",
                        "BloombergMarketSector": "",
                        "SettleDays": 0,
                        "PricingFactor": 1,
                        "CurrentPriceDayRange": 0,
                        "AssetType": "OP",
                        "InvestmentType": "CMFTOP",
                        "PriceDenomination": "",
                        "BifurcationCurrency": "INR",
                        "PrincipalCurrency": "INR",
                        "IncomeCurrency": "INR",
                        "RiskCurrency": "INR",
                        "IssueCountry": "IN",
                        "WithholdingTaxType": "Standard",
                        "QDIEligibilityFlag": "Not Eligible",
                        "SharesOutstanding": "",
                        "Beta": "",
                        "SubIndustry": "",
                        "SubIndustry2": "",
                        "NonDeliverableCurrencyFlag": "",
                        "QuantityPrecision": "",
                        "InvestmentCrossZero": "",
                        "PriceByPreference": "",
                        "ForwardPriceInterpolateFlag": "",
                        "SecFeeSchedule": "",
                        "SecEligibleFlag": "",
                        "WashSalesEligible": "",
                        "ExerciseStyle": "European",
                        "PricingPrecision": 15,
                        "CashSettledFlag": 1
                    },
                    "car_trade_loader": {
                        "RecordAction": "InsertUpdate",
                        "KeyValue.KeyName": "",
                        "UserTranId1": "",
                        "Portfolio": "AAFSPL_CAR",
                        "LocationAccount": "AAFSPL_CAR_Ventura Securities Ltd",
                        "Strategy": "Default",
                        "Broker": "",
                        "PriceDenomination": "CALC",
                        "CounterInvestment": "INR",
                        "NetInvestmentAmount": "CALC",
                        "NetCounterAmount": "CALC",
                        "TradeFX": "",
                        "ContractFxRateNumerator": "",
                        "ContractFxRateDenominator": "",
                        "ContractFxRate": "",
                        "NotionalAmount": "",
                        "FundStructure": "CALC",
                        "SpotDate": "",
                        "PriceDirectly": "",
                        "CounterFXDenomination": "CALC",
                        "CounterTDateFx": "",
                        "AccruedInterest": "",
                        "InvestmentAccruedInterest": "",
                        "TradeExpenses.ExpenseNumber": "",
                        "TradeExpenses.ExpenseCode": "",
                        "TradeExpenses.ExpenseAmt": "",
                        "NonCapExpenses.NonCapNumber": "",
                        "NonCapExpenses.NonCapExpenseCode": "",
                        "NonCapExpenses.NonCapAmount": "",
                        "NonCapExpenses.NonCapCurrency": "",
                        "NonCapExpenses.LocationAccount": "",
                        "NonCapExpenses.NonCapLiabilityCode": "",
                        "NonCapExpenses.NonCapPaymentType": ""
                    }
                }
                # Save the default consolidated data file
                with open(consolidated_path, "w") as f:
                    json.dump(consolidated_data, f, indent=4)
            
            lotsize_data = consolidated_data.get("lotsize_data", {})
            headers = consolidated_data.get("trade_headers", {})
            aafspl_car_future = consolidated_data.get("aafspl_car_future", {})
            option_security = consolidated_data.get("option_security", {})
            car_trade_loader = consolidated_data.get("car_trade_loader", {})
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load consolidated_data.json: {e}\nPlease ensure this file exists under my_app/.")
            return

        headers_values = list(headers.values())
        date_key = headers.get('DATE')
        scrip_key = headers.get('SCRIP_NAME')
        buy_qty_key = headers.get('BUY_QTY')
        buy_rate_key = headers.get('BUY_RATE')
        sell_qty_key = headers.get('SELL_QTY')
        sell_rate_key = headers.get('SELL_RATE')
        csv_path = self.csv_path_var.get().strip()
        if not csv_path or not os.path.exists(csv_path):
            messagebox.showwarning("CSV Missing", "Please select a valid Trade CSV file.")
            return

        try:
            df_csv = pd.read_csv(csv_path, usecols=headers_values)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read CSV: {e}")
            return
       
        for_trade, security_creation, aafspl_car_future_data, car_trade_loader_data, option_security_data = self._template_data(df_csv, date_key, scrip_key, buy_qty_key,
                            buy_rate_key,sell_qty_key,sell_rate_key,
                            lotsize_data, aafspl_car_future, option_security, car_trade_loader)

        # Persist in instance and render
        self._df_csv = df_csv
        self.for_trade_rows = for_trade
        self.security_creation_rows = list(security_creation.values())
        self._aafspl_car_future_data = aafspl_car_future_data
        self._car_trade_loader_data = car_trade_loader_data
        self._option_security_data = option_security_data

        self._render_for_trade()
        self._render_sec()

        self.status_var.set(f"Processed {len(self.for_trade_rows)} trades | {len(self.security_creation_rows)} securities")

    # ---- Rendering with filtering ----
    def _render_for_trade(self):
        # clear
        for it in self.for_trade_tree.get_children():
            self.for_trade_tree.delete(it)
        term = (self.for_trade_search_var.get() if hasattr(self, 'for_trade_search_var') else "").lower().strip()
        for r in self.for_trade_rows:
            if not term or term in " ".join(map(str, r)).lower():
                self.for_trade_tree.insert("", "end", values=r)

    def _render_sec(self):
        for it in self.sec_tree.get_children():
            self.sec_tree.delete(it)
        term = (self.sec_search_var.get() if hasattr(self, 'sec_search_var') else "").lower().strip()
        for r in self.security_creation_rows:
            if not term or term in " ".join(map(str, r)).lower():
                self.sec_tree.insert("", "end", values=r)

    def _export_excel(self):
        if not self.for_trade_rows and not self.security_creation_rows:
            messagebox.showinfo("Nothing to export", "Process a CSV first.")
            return

        # Ask output path
        out_path = filedialog.asksaveasfilename(
            title="Save Excel",
            defaultextension=".xlsx",
            filetypes=[["Excel", "*.xlsx"]],
            initialfile="TradeVentura_Output.xlsx"
        )
        if not out_path:
            return

        # Build DataFrames and write sheets like test.py output
        try:
            df_for_trade = pd.DataFrame(self.for_trade_rows, columns=["Scrip", "Buy/Sell", "Qty", "Rate"]) 
            df_sec = pd.DataFrame(self.security_creation_rows, columns=["Scrip Name", "Underlying", "Expiry", "Strike Price", "Call Put Flag", "Trading Size"]) 

            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                # Sheet 1: Original CSV (no bold headers, no borders)
                df_csv = getattr(self, "_df_csv", None)
                if df_csv is not None:
                    df_csv.to_excel(writer, sheet_name="Original_Data", index=False)
                else:
                    pd.DataFrame().to_excel(writer, sheet_name="Original_Data", index=False)
                # normalize header styling
                ws_orig = writer.book["Original_Data"]
                if ws_orig.max_row >= 1:
                    for cell in ws_orig[1]:
                        cell.font = Font(bold=False)
                        cell.border = Border()

                # Sheet 2: combined with merged section headers
                workbook = writer.book
                ws = workbook.create_sheet(title="Trade_and_Security")

                for_trade_header = ["Scrip", "Buy/Sell", "Qty", "Rate"]
                sec_creation_header = ["Scrip Name", "Underlying", "Expiry", "Strike Price", "Call Put Flag", "Trading Size"]

                # Merge headers
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(for_trade_header))
                ws.merge_cells(start_row=1, start_column=len(for_trade_header)+2, 
                               end_row=1, end_column=len(for_trade_header)+1+len(sec_creation_header))

                ws.cell(row=1, column=1, value="For_Trade").alignment = Alignment(horizontal="center")
                ws.cell(row=1, column=1).font = Font(bold=True)

                ws.cell(row=1, column=len(for_trade_header)+2, value="Security_Creation").alignment = Alignment(horizontal="center")
                ws.cell(row=1, column=len(for_trade_header)+2).font = Font(bold=True)

                # Sub headers
                for col, header in enumerate(for_trade_header, start=1):
                    ws.cell(row=2, column=col, value=header).font = Font(bold=True)
                for col, header in enumerate(sec_creation_header, start=len(for_trade_header)+2):
                    ws.cell(row=2, column=col, value=header).font = Font(bold=True)

                # Data rows
                max_len = max(len(self.for_trade_rows), len(self.security_creation_rows))
                sec_values = self.security_creation_rows
                for i in range(max_len):
                    if i < len(self.for_trade_rows):
                        for col, val in enumerate(self.for_trade_rows[i], start=1):
                            ws.cell(row=i+3, column=col, value=val)
                    if i < len(sec_values):
                        for col, val in enumerate(sec_values[i], start=len(for_trade_header)+2):
                            ws.cell(row=i+3, column=col, value=val)

            messagebox.showinfo("Success", f"Excel exported to:\n{out_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export excel: {e}")

    def _export_to_template(self):
        if not hasattr(self, '_aafspl_car_future_data') or not hasattr(self, '_car_trade_loader_data') or not hasattr(self, '_option_security_data'):
            messagebox.showinfo("Nothing to export", "Process a CSV first to generate template data.")
            return

        # Ask output path for zip file
        out_path = filedialog.asksaveasfilename(
            title="Save Template Data",
            defaultextension=".zip",
            filetypes=[["ZIP Files", "*.zip"]],
            initialfile="TradeVentura.zip"
        )
        if not out_path:
            return
        
        import time
        start_time = time.time()
        # 🔹 Show spinner (non-blocking)
        loader = LoadingSpinner(self, text="Exporting templates...")

        def task():

            try:
                # Create Excel files using helper functions
                excel_files = []
                
                # Create aafspl_car_future Excel
                if self._aafspl_car_future_data:
                    # print(f"Creating aafspl Excel with {len(self._aafspl_car_future_data)} rows...")
                    excel_start = time.time()
                    aafspl_excel, aafspl_name = output_save_in_template(
                        self._aafspl_car_future_data, 
                        AAFSPL_HEADER, 
                        "aafspl_car_future_data.xlsx"
                    )
                    # print(f"aafspl Excel created in {time.time() - excel_start:.4f} seconds")
                    excel_files.append((aafspl_excel, aafspl_name))
                
                # Create car_trade_loader Excel
                if self._car_trade_loader_data:
                    # print(f"Creating car_trade Excel with {len(self._car_trade_loader_data)} rows...")
                    excel_start = time.time()
                    car_trade_excel, car_trade_name = output_save_in_template(
                        self._car_trade_loader_data, 
                        CAR_TRADE_HEADERS, 
                        "car_trade_loader_data.xlsx"
                    )
                    # print(f"car_trade Excel created in {time.time() - excel_start:.4f} seconds")
                    excel_files.append((car_trade_excel, car_trade_name))
                
                # Create option_security Excel
                if self._option_security_data:
                    # print(f"Creating option Excel with {len(self._option_security_data)} rows...")
                    excel_start = time.time()
                    option_excel, option_name = output_save_in_template(
                        self._option_security_data, 
                        OPTION_HEADER, 
                        "option_security_data.xlsx"
                    )
                    # print(f"option Excel created in {time.time() - excel_start:.4f} seconds")
                    excel_files.append((option_excel, option_name))
                
                # Create zip file
                zip_buffer = multiple_excels_to_zip(excel_files, "TradeVentura.zip")
                
                # Save zip file
                with open(out_path, 'wb') as f:
                    f.write(zip_buffer.read())
                
                file_list = [name for _, name in excel_files]
                # ✅ Close loader + show success
                loader.close()
                messagebox.showinfo("Success", f"Template data exported to:\n{out_path}\n\nContains:\n" + "\n".join([f"- {file}" for file in file_list]))
            except Exception as e:
                loader.close()
                messagebox.showerror("Error", f"Failed to export template data: {e}")
            finally:
                end_time = time.time()
                # print(f"Time taken: {end_time - start_time:.4f} seconds")
        
        # 🔹 Run heavy work in thread
        threading.Thread(target=task, daemon=True).start()


    def _template_data(self, df_csv, date_key, scrip_key, buy_qty_key,
                            buy_rate_key,sell_qty_key,sell_rate_key,
                            lotsize_data, aafspl_car_future, option_security, car_trade_loader):
        
        # Build rows following test.py logic
        for_trade = []
        security_creation = {}
        aafspl_car_future_data = []
        car_trade_loader_data = []
        option_security_data = []

        # Pre-build template dictionaries for performance (only once)
        option_template = {
            OPTION_CODE                         : "",
            OPTION_KEYVALUE                     : "",
            OPTION_DESCRIPTION                  : "",
            OPTION_EXTENDEDDESCRIPTION          : "",
            OPTION_EXCHANGE                     : option_security[OPTION_EXCHANGE],
            OPTION_ISSUER                       : option_security[OPTION_ISSUER],
            OPTION_TICKER                       : option_security[OPTION_TICKER],
            OPTION_CUSIP                        : option_security[OPTION_CUSIP],
            OPTION_SEDOL                        : option_security[OPTION_SEDOL],
            OPTION_ISIN                         : option_security[OPTION_ISIN],
            OPTION_RICCODE                      : option_security[OPTION_RICCODE],
            OPTION_ALTKEY1                      : option_security[OPTION_ALTKEY1],
            OPTION_ALTKEY2                      : option_security[OPTION_ALTKEY2],
            OPTION_BLOOMBERGID                  : option_security[OPTION_BLOOMBERGID],
            OPTION_BLOOMBERGTICKER              : option_security[OPTION_BLOOMBERGTICKER],
            OPTION_BLOOMBERGUNIQUEID            : option_security[OPTION_BLOOMBERGUNIQUEID],
            OPTION_BLOOMBERGMARKETSECTOR        : option_security[OPTION_BLOOMBERGMARKETSECTOR],
            OPTION_SETTLEDAYS                   : option_security[OPTION_SETTLEDAYS],
            OPTION_PRICINGFACTOR                : option_security[OPTION_PRICINGFACTOR],
            OPTION_TRADINGFACTOR                : 0,
            OPTION_CURRENTPRICEDAYRANGE         : option_security[OPTION_CURRENTPRICEDAYRANGE],
            OPTION_ASSETTYPE                    : option_security[OPTION_ASSETTYPE],
            OPTION_INVESTMENTTYPE               : option_security[OPTION_INVESTMENTTYPE],
            OPTION_PRICEDENOMINATION            : option_security[OPTION_PRICEDENOMINATION],
            OPTION_BIFURCATIONCURRENCY          : option_security[OPTION_BIFURCATIONCURRENCY],
            OPTION_PRINCIPALCURRENCY            : option_security[OPTION_PRINCIPALCURRENCY],
            OPTION_INCOMECURRENCY               : option_security[OPTION_INCOMECURRENCY],
            OPTION_RISKCURRENCY                 : option_security[OPTION_RISKCURRENCY],
            OPTION_ISSUECOUNTRY                 : option_security[OPTION_ISSUECOUNTRY],
            OPTION_WITHHOLDINGTAXTYPE           : option_security[OPTION_WITHHOLDINGTAXTYPE],
            OPTION_QDIELIGIBILITYFLAG           : option_security[OPTION_QDIELIGIBILITYFLAG],
            OPTION_SHARESOUTSTANDING            : option_security[OPTION_SHARESOUTSTANDING],
            OPTION_BETA                         : option_security[OPTION_BETA],
            OPTION_SUBINDUSTRY                  : option_security[OPTION_SUBINDUSTRY],
            OPTION_SUBINDUSTRY2                 : option_security[OPTION_SUBINDUSTRY2],
            OPTION_NONDELIVERABLECURRENCYFLAG   : option_security[OPTION_NONDELIVERABLECURRENCYFLAG],
            OPTION_QUANTITYPRECISION            : option_security[OPTION_QUANTITYPRECISION],
            OPTION_INVESTMENTCROSSZERO          : option_security[OPTION_INVESTMENTCROSSZERO],
            OPTION_UNDERLYINGINVESTMENT         : "",
            OPTION_EXPIREDATE                   : "",
            OPTION_STRIKEPRICE                  : "",
            OPTION_PRICEBYPREFERENCE            : option_security[OPTION_PRICEBYPREFERENCE],
            OPTION_FORWARDPRICEINTERPOLATEFLAG  : option_security[OPTION_FORWARDPRICEINTERPOLATEFLAG],
            OPTION_SECFEESCHEDULE               : option_security[OPTION_SECFEESCHEDULE],
            OPTION_SECELIGIBLEFLAG              : option_security[OPTION_SECELIGIBLEFLAG],
            OPTION_WASHSALESELIGIBLE            : option_security[OPTION_WASHSALESELIGIBLE],
            OPTION_EXERCISESTYLE                : option_security[OPTION_EXERCISESTYLE],
            OPTION_PUTCALLFLAG                  : 0,
            OPTION_PRICINGPRECISION             : option_security[OPTION_PRICINGPRECISION],
            OPTION_CASHSETTLEDFLAG              : option_security[OPTION_CASHSETTLEDFLAG],
        }

        car_trade_template = {
            CAR_TRADE_RECORDTYPE                        : "",
            CAR_TRADE_RECORDACTION                      : car_trade_loader[CAR_TRADE_RECORDACTION],
            CAR_TRADE_KEYVALUE2                         : "",
            CAR_TRADE_KEYVALUE                          : car_trade_loader[CAR_TRADE_KEYVALUE],
            CAR_TRADE_USERTRANID1                       : car_trade_loader[CAR_TRADE_USERTRANID1],
            CAR_TRADE_PORTFOLIO                         : car_trade_loader[CAR_TRADE_PORTFOLIO],
            CAR_TRADE_LOCATIONACCOUNT                   : car_trade_loader[CAR_TRADE_LOCATIONACCOUNT],
            CAR_TRADE_STRATEGY                          : car_trade_loader[CAR_TRADE_STRATEGY],
            CAR_TRADE_INVESTMENT                        : "",
            CAR_TRADE_BROKER                            : car_trade_loader[CAR_TRADE_BROKER],
            CAR_TRADE_EVENTDATE                         : "",
            CAR_TRADE_SETTLEDATE                        : "",
            CAR_TRADE_ACTUALSETTLEDATE                  : "",
            CAR_TRADE_QUANTITY                          : 0,
            CAR_TRADE_PRICE                             : 0,
            CAR_TRADE_PRICEDENOMINATION                 : car_trade_loader[CAR_TRADE_PRICEDENOMINATION],
            CAR_TRADE_COUNTERINVESTMENT                 : car_trade_loader[CAR_TRADE_COUNTERINVESTMENT],
            CAR_TRADE_NETINVESTMENTAMOUNT               : car_trade_loader[CAR_TRADE_NETINVESTMENTAMOUNT],
            CAR_TRADE_NETCOUNTERAMOUNT                  : car_trade_loader[CAR_TRADE_NETCOUNTERAMOUNT],
            CAR_TRADE_TRADEFX                           : car_trade_loader[CAR_TRADE_TRADEFX],
            CAR_TRADE_CONTRACTFXRATENUMERATOR           : car_trade_loader[CAR_TRADE_CONTRACTFXRATENUMERATOR],
            CAR_TRADE_CONTRACTFXRATEDENOMINATOR         : car_trade_loader[CAR_TRADE_CONTRACTFXRATEDENOMINATOR],
            CAR_TRADE_CONTRACTFXRATE                    : car_trade_loader[CAR_TRADE_CONTRACTFXRATE],
            CAR_TRADE_NOTIONALAMOUNT                    : car_trade_loader[CAR_TRADE_NOTIONALAMOUNT],
            CAR_TRADE_FUNDSTRUCTURE                     : car_trade_loader[CAR_TRADE_FUNDSTRUCTURE],
            CAR_TRADE_SPOTDATE                          : car_trade_loader[CAR_TRADE_SPOTDATE],
            CAR_TRADE_PRICEDIRECTLY                     : car_trade_loader[CAR_TRADE_PRICEDIRECTLY],
            CAR_TRADE_COUNTERFXDENOMINATION             : car_trade_loader[CAR_TRADE_COUNTERFXDENOMINATION],
            CAR_TRADE_COUNTERTDATEFX                    : car_trade_loader[CAR_TRADE_COUNTERTDATEFX],
            CAR_TRADE_ACCRUEDINTEREST                   : car_trade_loader[CAR_TRADE_ACCRUEDINTEREST],
            CAR_TRADE_INVESTMENTACCRUEDINTEREST         : car_trade_loader[CAR_TRADE_INVESTMENTACCRUEDINTEREST],
            CAR_TRADE_TRADEEXPENSES_EXPENSENUMBER       : car_trade_loader[CAR_TRADE_TRADEEXPENSES_EXPENSENUMBER],
            CAR_TRADE_TRADEEXPENSES_EXPENSECODE         : car_trade_loader[CAR_TRADE_TRADEEXPENSES_EXPENSECODE],
            CAR_TRADE_TRADEEXPENSES_EXPENSEAMT          : car_trade_loader[CAR_TRADE_TRADEEXPENSES_EXPENSEAMT],
            CAR_TRADE_NONCAPEXPENSES_NONCAPNUMBER       : car_trade_loader[CAR_TRADE_NONCAPEXPENSES_NONCAPNUMBER],
            CAR_TRADE_NONCAPEXPENSES_NONCAPEXPENSECODE  : car_trade_loader[CAR_TRADE_NONCAPEXPENSES_NONCAPEXPENSECODE],
            CAR_TRADE_NONCAPEXPENSES_NONCAPAMOUNT       : car_trade_loader[CAR_TRADE_NONCAPEXPENSES_NONCAPAMOUNT],
            CAR_TRADE_NONCAPEXPENSES_NONCAPCURRENCY     : car_trade_loader[CAR_TRADE_NONCAPEXPENSES_NONCAPCURRENCY],
            CAR_TRADE_NONCAPEXPENSES_LOCATIONACCOUNT    : car_trade_loader[CAR_TRADE_NONCAPEXPENSES_LOCATIONACCOUNT],
            CAR_TRADE_NONCAPEXPENSES_NONCAPLIABILITYCODE: car_trade_loader[CAR_TRADE_NONCAPEXPENSES_NONCAPLIABILITYCODE],
            CAR_TRADE_NONCAPEXPENSES_NONCAPPAYMENTTYPE  : car_trade_loader[CAR_TRADE_NONCAPEXPENSES_NONCAPPAYMENTTYPE]
        }

        aafspl_template = {
            AAFSPL_CODE :      "",
            AAFSPL_KEYVALUE :  "",
            AAFSPL_DESCRIPTION :  "",
            AAFSPL_EXTENDEDDESCRIPTION :  "",
            AAFSPL_EXCHANGE : aafspl_car_future[AAFSPL_EXCHANGE],
            AAFSPL_ISSUER : aafspl_car_future[AAFSPL_ISSUER],
            AAFSPL_TICKER : aafspl_car_future[AAFSPL_TICKER],
            AAFSPL_CUSIP : aafspl_car_future[AAFSPL_CUSIP],
            AAFSPL_SEDOL : aafspl_car_future[AAFSPL_SEDOL],
            AAFSPL_ISIN : aafspl_car_future[AAFSPL_ISIN],
            AAFSPL_ALTKEY1 : aafspl_car_future[AAFSPL_ALTKEY1],
            AAFSPL_ALTKEY2 : aafspl_car_future[AAFSPL_ALTKEY2],
            AAFSPL_BLOOMBERGID : aafspl_car_future[AAFSPL_BLOOMBERGID],
            AAFSPL_BLOOMBERGTICKER : aafspl_car_future[AAFSPL_BLOOMBERGTICKER],
            AAFSPL_BLOOMBERGUNIQUEID : aafspl_car_future[AAFSPL_BLOOMBERGUNIQUEID],
            AAFSPL_BLOOMBERGMARKETSECTOR : aafspl_car_future[AAFSPL_BLOOMBERGMARKETSECTOR],
            AAFSPL_SETTLEDAYS : aafspl_car_future[AAFSPL_SETTLEDAYS],
            AAFSPL_PRICINGFACTOR : aafspl_car_future[AAFSPL_PRICINGFACTOR],
            AAFSPL_TRADINGFACTOR : 0,
            AAFSPL_CURRENTPRICEDAYRANGE : aafspl_car_future[AAFSPL_CURRENTPRICEDAYRANGE],
            AAFSPL_ASSETTYPE : aafspl_car_future[AAFSPL_ASSETTYPE],
            AAFSPL_INVESTMENTTYPE : aafspl_car_future[AAFSPL_INVESTMENTTYPE],
            AAFSPL_PRICEDENOMINATION : aafspl_car_future[AAFSPL_PRICEDENOMINATION],
            AAFSPL_BIFURCATIONCURRENCY : aafspl_car_future[AAFSPL_BIFURCATIONCURRENCY],
            AAFSPL_PRINCIPALCURRENCY : aafspl_car_future[AAFSPL_PRINCIPALCURRENCY],
            AAFSPL_INCOMECURRENCY : aafspl_car_future[AAFSPL_INCOMECURRENCY],
            AAFSPL_RISKCURRENCY : aafspl_car_future[AAFSPL_RISKCURRENCY],
            AAFSPL_ISSUECOUNTRY : aafspl_car_future[AAFSPL_ISSUECOUNTRY],
            AAFSPL_WITHHOLDINGTAXTYPE : aafspl_car_future[AAFSPL_WITHHOLDINGTAXTYPE],
            AAFSPL_QDIELIGIBILITYFLAG : aafspl_car_future[AAFSPL_QDIELIGIBILITYFLAG],
            AAFSPL_SHARESOUTSTANDING : aafspl_car_future[AAFSPL_SHARESOUTSTANDING],
            AAFSPL_SUBINDUSTRY : aafspl_car_future[AAFSPL_SUBINDUSTRY],
            AAFSPL_SUBINDUSTRY2 : aafspl_car_future[AAFSPL_SUBINDUSTRY2],
            AAFSPL_QUANTITYPRECISION : aafspl_car_future[AAFSPL_QUANTITYPRECISION],
            AAFSPL_INVESTMENTCROSSZERO : aafspl_car_future[AAFSPL_INVESTMENTCROSSZERO],
            AAFSPL_UNDERLYINGINVESTMENT : "",
            AAFSPL_EXPIREDATE : "",
            AAFSPL_STRIKEPRICE : "",
            AAFSPL_PRICEBYPREFERENCE : aafspl_car_future[AAFSPL_PRICEBYPREFERENCE],
            AAFSPL_FORWARDPRICEINTERPOLATEFLAG : aafspl_car_future[AAFSPL_FORWARDPRICEINTERPOLATEFLAG],
            AAFSPL_PRICINGPRECISION : aafspl_car_future[AAFSPL_PRICINGPRECISION],
            AAFSPL_FIRSTMARKDATE : aafspl_car_future[AAFSPL_FIRSTMARKDATE],
            AAFSPL_LASTMARKDATE : aafspl_car_future[AAFSPL_LASTMARKDATE],
            AAFSPL_PRICELIST : aafspl_car_future[AAFSPL_PRICELIST],
            AAFSPL_AUTOGENERATEMARKS : aafspl_car_future[AAFSPL_AUTOGENERATEMARKS],
            AAFSPL_CASHSETTLEMENT : aafspl_car_future[AAFSPL_CASHSETTLEMENT]
        }

        def genevascrip(scrip: str):
            parts = scrip.split()
            yyyymmdd, mm_dd_yyyy = _format_date(parts[1])
            if len(parts) == 2:
                parts += ["F", "0"]
            underlying = parts[0]
            expiry_date = mm_dd_yyyy + ' 23:59:59'
            # strike_price = parts[3]
            strike_price = int(float(parts[3])) if float(parts[3]).is_integer() else float(parts[3])
            call_or_put_flag = 0 if parts[2] == "CE" else 1
            trading_size = lotsize_data.get(parts[0], 0)
            return underlying, expiry_date, strike_price, call_or_put_flag, trading_size, f"MCX{parts[0]}{yyyymmdd}{parts[2][0]}{parts[3]}"

        # Process rows using faster iteration
        import time
        start_time = time.time()
        # print(f"Starting to process {len(df_csv)} rows...")
        
        # Convert to list for faster iteration (much faster than iterrows)
        rows_data = df_csv.to_dict('records')
        
        # Track unique scrip names for filtering
        seen_option_scrips = set()
        seen_aafspl_scrips = set()
        
        for idx, row in enumerate(rows_data):            # Clean string values
            for key, value in row.items():
                if isinstance(value, str):
                    row[key] = value.strip()

            date = row.get(date_key)
            date_obj = datetime.strptime(date, "%d-%b-%y")
            formatted_date = date_obj.strftime("%d/%m/%Y")

            buy_qty = _safe_decimal(row.get(buy_qty_key))
            sell_qty = _safe_decimal(row.get(sell_qty_key))
            buy_rate = _safe_decimal(row.get(buy_rate_key))
            sell_rate = _safe_decimal(row.get(sell_rate_key))

            buy_sell = "Sell" if buy_qty == 0 else "Buy"
            abs_diff = abs(buy_qty - sell_qty)
            abs_rate_diff = abs(buy_rate - sell_rate)

            try:
                underlying, expiry_date, strike_price, call_or_put_flag, trading_size, scrip_name = genevascrip(row.get(scrip_key))
            except Exception as e:
                # Skip malformed scrip lines
                continue

            for_trade.append([scrip_name, buy_sell, str(abs_diff), str(abs_rate_diff)])
            security_creation[scrip_name] = [
                scrip_name,
                underlying,
                expiry_date,
                strike_price,
                call_or_put_flag,
                trading_size,
            ]

            # Parse scrip to get option type (CE/PE)
            scrip_parts = row.get(scrip_key).split()
            option_type = scrip_parts[2] if len(scrip_parts) > 2 else "F"
            # Filter option_record: Only unique scrip_name and only PE/CE options
            if option_type in ["CE", "PE"] and scrip_name not in seen_option_scrips:
                seen_option_scrips.add(scrip_name)
                # Use template copy for performance (much faster than creating new dict)
                option_record = option_template.copy()
                option_record.update({
                    OPTION_CODE: scrip_name,
                    OPTION_KEYVALUE: scrip_name,
                    OPTION_DESCRIPTION: scrip_name,
                    OPTION_EXTENDEDDESCRIPTION: scrip_name,
                    OPTION_TRADINGFACTOR: trading_size,
                    OPTION_UNDERLYINGINVESTMENT: underlying,
                    OPTION_EXPIREDATE: expiry_date,
                    OPTION_STRIKEPRICE: strike_price,
                    OPTION_PUTCALLFLAG: call_or_put_flag,
                })
                option_security_data.append(option_record)
            
            # Use template copy for performance
            car_trade_record = car_trade_template.copy()
            car_trade_record.update({
                CAR_TRADE_RECORDTYPE: buy_sell,
                CAR_TRADE_KEYVALUE2: scrip_name,
                CAR_TRADE_INVESTMENT: scrip_name,
                CAR_TRADE_EVENTDATE: formatted_date,
                CAR_TRADE_SETTLEDATE: formatted_date,
                CAR_TRADE_ACTUALSETTLEDATE: formatted_date,
                CAR_TRADE_QUANTITY: abs_diff,
                CAR_TRADE_PRICE: abs_rate_diff,
            })
            car_trade_loader_data.append(car_trade_record)

            # Filter aafspl_record: Only F (Future) and unique scrip_name
            if option_type == "F" and scrip_name not in seen_aafspl_scrips:
                seen_aafspl_scrips.add(scrip_name)
                # Use template copy for performance
                aafspl_record = aafspl_template.copy()
                aafspl_record.update({
                    AAFSPL_CODE: scrip_name,
                    AAFSPL_KEYVALUE: scrip_name,
                    AAFSPL_DESCRIPTION: scrip_name,
                    AAFSPL_EXTENDEDDESCRIPTION: scrip_name,
                    AAFSPL_TRADINGFACTOR: trading_size,
                    AAFSPL_UNDERLYINGINVESTMENT: underlying,
                    AAFSPL_EXPIREDATE: expiry_date,
                    AAFSPL_STRIKEPRICE: strike_price,
                })
                aafspl_car_future_data.append(aafspl_record)
        
        end_time = time.time()
        # print(f"Data processing took: {end_time - start_time:.4f} seconds")
        
        # Debug information for filtering
        # print(f"Filtering Results:")
        # print(f"- Total option records (CE/PE only, unique): {len(option_security_data)}")
        # print(f"- Total aafspl records (F only, unique): {len(aafspl_car_future_data)}")
        # print(f"- Total car_trade records: {len(car_trade_loader_data)}")
        # print(f"- Unique option scrips seen: {len(seen_option_scrips)}")
        # print(f"- Unique aafspl scrips seen: {len(seen_aafspl_scrips)}")
            
        return for_trade, security_creation, aafspl_car_future_data, car_trade_loader_data,option_security_data
