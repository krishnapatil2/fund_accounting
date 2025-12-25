from decimal import Decimal
import decimal
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import json
from datetime import datetime
import pandas as pd
from my_app.CONSTANTS import CDS_GENEVA_HEADER_LIST, REG_GENEVA_HEADER_LIST
from my_app.pages.helper import parse_expiry_date, read_file
from my_app.pages.loading import LoadingSpinner
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io
import zipfile
import threading
from collections import defaultdict

# Header configurations for F&O files
CDS_HOLDINGS_HEADER = ['Source.Name','Date','QuantityUnit','Exchange','ClientCode','TradingCode','UnderlyingCode','ClientName','UnderlyingName','InstrumentType','ExpiryDate','OptionType','StrikePrice','OpenBuy','OpenSell','TradedBuy','DayBuyValue','TradedSell','DaySellValue','ExcerciseQty','AllocationQty','NetBuy','NetSell','ContractSettlementPrice','Settlement Price','BloombergCodes','Concatenate','Fund Name In Geneva','UniqueCode','Netbuy-Netsell']

REGULAR_HOLDINGS_HEADERS = [
    'Source.Name', 'Date', 'Exchange', 'ClientName', 'ClientCode', 'OptionType',
    'UnderlyingCode', 'UnderlyingName', 'InstrumentType', 'StrikePrice', 'ExpiryDate',
    'OpenBuy', 'OpenSell', 'TradedSell', 'TradedBuy', 'DayBuyValue', 'DaySellValue',
    'ExcerciseQty', 'AllocationQty', 'NetBuy', 'NetSell', 'ContractSettlementPrice',
    'ClosingPrice', 'BloombergCD', 'Concatenate', 'Fund Name In Geneva', 'UniqueCode',
    'NerBuy-Netsell'
]

GENEVA_HOLDINGS_HEADERS = [
    'Portfolio', 'Investment', 'Investment Description', 'Traded Quantity', 'Settled Quantity',
    'Currency', 'Unit Cost', 'Cost Local', 'Cost Book', 'Market Price Local', 'Market Value Local',
    'Market Value Book', 'Unrealized G/L Book', 'Accrued Interest Local', 'Cust Account',
    'Strategy', 'UniqueCode', 'Fund Name IN Custody Holding', 'Nse Price', 'Diff'
]


NSE_F_AND_O_BHAVCOPY = [
    'INSTRUMENT', 'SYMBOL', 'EXPIRY_DT', 'STRIKE_PR', 'OPTION_TYP', 'OPEN',
    'HIGH', 'LOW', 'CLOSE', 'SETTLE_PR', 'CONTRACTS', 'VAL_INLAKH',
    'OPEN_INT', 'CHG_IN_OI', 'TIMESTAMP', 'ConcatenateCode'
]


class FOReconciliationPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg="#ecf0f1")

        # Title
        title = tk.Label(self, text="📈 Daily F&O Reconciliation", font=("Arial", 20, "bold"), bg="#ecf0f1", fg="#2c3e50")
        title.pack(pady=10)

        # Main content frame
        content_frame = tk.Frame(self, bg="#ecf0f1")
        content_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Left panel - Holding Statements Raw Files Browser
        left_panel = tk.Frame(content_frame, bg="#ecf0f1")
        left_panel.pack(side="left", fill="both", expand=True, padx=(0, 10))

        # Holding Statements Raw Files Browser Section
        holdings_frame = tk.LabelFrame(left_panel, text="📁 Holding Statements Raw Files", font=("Arial", 12, "bold"), 
                                      bg="#ecf0f1", fg="#2c3e50", padx=10, pady=10)
        holdings_frame.pack(fill="both", expand=True)

        # Browse buttons frame
        holdings_buttons_frame = tk.Frame(holdings_frame, bg="#ecf0f1")
        holdings_buttons_frame.pack(fill="x", pady=(0, 10))

        # Browse Files button
        self.browse_holdings_btn = tk.Button(holdings_buttons_frame, text="📂 Browse Raw Files", 
                                            bg="#3498db", fg="white", font=("Arial", 11, "bold"), 
                                            relief="flat", padx=15, pady=8, command=self.browse_holdings_files)
        self.browse_holdings_btn.pack(side="left", padx=(0, 10))

        # Clear Files button
        self.clear_holdings_btn = tk.Button(holdings_buttons_frame, text="🗑️ Clear All", 
                                           bg="#e74c3c", fg="white", font=("Arial", 11, "bold"), 
                                           relief="flat", padx=15, pady=8, command=self.clear_holdings_files)
        self.clear_holdings_btn.pack(side="left")

        # Files listbox with scrollbar
        holdings_listbox_frame = tk.Frame(holdings_frame, bg="#ecf0f1")
        holdings_listbox_frame.pack(fill="both", expand=True)

        # Listbox for displaying holding files
        self.holdings_files_listbox = tk.Listbox(holdings_listbox_frame, font=("Arial", 10), bg="white", 
                                                selectmode=tk.EXTENDED, relief="solid", bd=1)
        holdings_scrollbar = tk.Scrollbar(holdings_listbox_frame, orient="vertical", command=self.holdings_files_listbox.yview)
        self.holdings_files_listbox.configure(yscrollcommand=holdings_scrollbar.set)

        self.holdings_files_listbox.pack(side="left", fill="both", expand=True)
        holdings_scrollbar.pack(side="right", fill="y")

        # File info frame
        holdings_info_frame = tk.Frame(holdings_frame, bg="#ecf0f1")
        holdings_info_frame.pack(fill="x", pady=(10, 0))

        self.holdings_count_label = tk.Label(holdings_info_frame, text="No raw files selected", 
                                            font=("Arial", 10), bg="#ecf0f1", fg="#7f8c8d")
        self.holdings_count_label.pack(side="left")

        # Right panel - Geneva & BhavCopy File Browser
        right_panel = tk.Frame(content_frame, bg="#ecf0f1")
        right_panel.pack(side="right", fill="both", expand=True)

        # Geneva File Browser Section
        geneva_frame = tk.LabelFrame(right_panel, text="🏦 Geneva Holdings File", font=("Arial", 12, "bold"), 
                                   bg="#ecf0f1", fg="#2c3e50", padx=10, pady=10)
        geneva_frame.pack(fill="both", expand=True, pady=(0, 10))

        # Geneva file path frame
        geneva_path_frame = tk.Frame(geneva_frame, bg="#ecf0f1")
        geneva_path_frame.pack(fill="x", pady=(0, 10))

        tk.Label(geneva_path_frame, text="Geneva File:", font=("Arial", 11), bg="#ecf0f1", fg="#2c3e50").pack(anchor="w")
        
        geneva_input_frame = tk.Frame(geneva_path_frame, bg="#ecf0f1")
        geneva_input_frame.pack(fill="x", pady=(5, 0))

        self.geneva_path_var = tk.StringVar()
        self.geneva_path_entry = tk.Entry(geneva_input_frame, textvariable=self.geneva_path_var, 
                                        font=("Arial", 10), width=50, relief="solid", bd=1)
        self.geneva_path_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        self.browse_geneva_btn = tk.Button(geneva_input_frame, text="Browse", 
                                         bg="#9b59b6", fg="white", font=("Arial", 10, "bold"), 
                                         relief="flat", padx=12, pady=4, command=self.browse_geneva_file)
        self.browse_geneva_btn.pack(side="right")

        # Geneva file info
        self.geneva_info_label = tk.Label(geneva_frame, text="No Geneva file selected", 
                                        font=("Arial", 10), bg="#ecf0f1", fg="#7f8c8d")
        self.geneva_info_label.pack(anchor="w", pady=(10, 0))

        # BhavCopy File Browser Section
        bhavcopy_frame = tk.LabelFrame(right_panel, text="📊 NSE F&O BhavCopy File", font=("Arial", 12, "bold"), 
                                      bg="#ecf0f1", fg="#2c3e50", padx=10, pady=10)
        bhavcopy_frame.pack(fill="both", expand=True)

        # BhavCopy file path frame
        bhavcopy_path_frame = tk.Frame(bhavcopy_frame, bg="#ecf0f1")
        bhavcopy_path_frame.pack(fill="x", pady=(0, 10))

        tk.Label(bhavcopy_path_frame, text="BhavCopy File:", font=("Arial", 11), bg="#ecf0f1", fg="#2c3e50").pack(anchor="w")
        
        bhavcopy_input_frame = tk.Frame(bhavcopy_path_frame, bg="#ecf0f1")
        bhavcopy_input_frame.pack(fill="x", pady=(5, 0))

        self.bhavcopy_path_var = tk.StringVar()
        self.bhavcopy_path_entry = tk.Entry(bhavcopy_input_frame, textvariable=self.bhavcopy_path_var, 
                                          font=("Arial", 10), width=50, relief="solid", bd=1)
        self.bhavcopy_path_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        self.browse_bhavcopy_btn = tk.Button(bhavcopy_input_frame, text="Browse", 
                                           bg="#9b59b6", fg="white", font=("Arial", 10, "bold"), 
                                           relief="flat", padx=12, pady=4, command=self.browse_bhavcopy_file)
        self.browse_bhavcopy_btn.pack(side="right")

        # BhavCopy file info
        self.bhavcopy_info_label = tk.Label(bhavcopy_frame, text="No BhavCopy file selected", 
                                          font=("Arial", 10), bg="#ecf0f1", fg="#7f8c8d")
        self.bhavcopy_info_label.pack(anchor="w", pady=(10, 0))

        # Process buttons frame
        process_frame = tk.Frame(self, bg="#ecf0f1")
        process_frame.pack(fill="x", padx=20, pady=20)

        # Process button
        self.process_btn = tk.Button(process_frame, text="🔄 Start Reconciliation", 
                                   bg="#27ae60", fg="white", font=("Arial", 12, "bold"), 
                                   relief="flat", padx=20, pady=10, command=self.start_reconciliation)
        self.process_btn.pack(side="left", padx=(0, 10))

        # Export button
        self.export_btn = tk.Button(process_frame, text="📊 Export Results", 
                                  bg="#8e44ad", fg="white", font=("Arial", 12, "bold"), 
                                  relief="flat", padx=20, pady=10, command=self.export_results,
                                  state="disabled")
        self.export_btn.pack(side="left")

        # Status bar
        self.status_var = tk.StringVar(value="Ready - Select all files to begin F&O reconciliation")
        status_label = tk.Label(self, textvariable=self.status_var, font=("Arial", 10), 
                              bg="#ecf0f1", fg="#7f8c8d", anchor="w")
        status_label.pack(fill="x", padx=20, pady=(0, 10))

        # Data storage
        self.holdings_files = []
        self.geneva_file_path = ""
        self.bhavcopy_file_path = ""
        self.reconciliation_results = None
        self.geneva_data = None
        self.bhavcopy_data = None
        self.bhavcopy_price_dict = {}  # Dictionary for concatenated key and closing price
        self.Reg_Geneva = []
        self.Cds_Geneva = []

    def browse_holdings_files(self):
        """Browse and select Holding Statements Raw files"""
        file_types = [
            ("All Supported Files", "*.xlsx *.xls *.csv"),
            ("Excel Files", "*.xlsx *.xls"),
            ("CSV Files", "*.csv"),
            ("All Files", "*.*")
        ]
        
        files = filedialog.askopenfilenames(
            title="Select Holding Statements Raw Files",
            filetypes=file_types
        )
        
        if files:
            self.holdings_files = list(files)
            self.update_holdings_listbox()
            self.update_holdings_count()
            self.status_var.set(f"Selected {len(self.holdings_files)} raw files")

    def browse_geneva_file(self):
        """Browse for Geneva Holdings file"""
        file_types = [
            ("All Supported Files", "*.xlsx *.xls *.csv"),
            ("Excel Files", "*.xlsx *.xls"),
            ("CSV Files", "*.csv"),
            ("All Files", "*.*")
        ]
        
        file_path = filedialog.askopenfilename(
            title="Select Geneva Holdings File",
            filetypes=file_types
        )
        
        if file_path:
            self.geneva_file_path = file_path
            self.geneva_path_var.set(file_path)
            file_name = os.path.basename(file_path)
            self.geneva_info_label.config(
                text=f"Selected: {file_name}",
                fg="#27ae60"
            )
            self.status_var.set(f"Geneva file selected: {file_name}")

    def browse_bhavcopy_file(self):
        """Browse for NSE F&O BhavCopy file"""
        file_types = [
            ("All Supported Files", "*.xlsx *.xls *.csv"),
            ("Excel Files", "*.xlsx *.xls"),
            ("CSV Files", "*.csv"),
            ("All Files", "*.*")
        ]
        
        file_path = filedialog.askopenfilename(
            title="Select NSE F&O BhavCopy File",
            filetypes=file_types
        )
        
        if file_path:
            self.bhavcopy_file_path = file_path
            self.bhavcopy_path_var.set(file_path)
            file_name = os.path.basename(file_path)
            self.bhavcopy_info_label.config(
                text=f"Selected: {file_name}",
                fg="#27ae60"
            )
            self.status_var.set(f"BhavCopy file selected: {file_name}")

    def clear_holdings_files(self):
        """Clear all selected holding files"""
        self.holdings_files = []
        self.update_holdings_listbox()
        self.update_holdings_count()
        self.status_var.set("Holding files cleared")

    def update_holdings_listbox(self):
        """Update the holdings files listbox"""
        self.holdings_files_listbox.delete(0, tk.END)
        for file_path in self.holdings_files:
            file_name = os.path.basename(file_path)
            self.holdings_files_listbox.insert(tk.END, file_name)

    def update_holdings_count(self):
        """Update the holdings file count label"""
        count = len(self.holdings_files)
        if count == 0:
            self.holdings_count_label.config(text="No raw files selected", fg="#7f8c8d")
        else:
            self.holdings_count_label.config(text=f"{count} file(s) selected", fg="#27ae60")

    def start_reconciliation(self):
        """Start the F&O reconciliation process"""
        # Validate inputs
        if not self.holdings_files:
            messagebox.showwarning("No Files", "Please select Holding Statements Raw files.")
            return
        
        if not self.geneva_file_path:
            messagebox.showwarning("No Geneva File", "Please select a Geneva Holdings file.")
            return
        
        if not self.bhavcopy_file_path:
            messagebox.showwarning("No BhavCopy File", "Please select a NSE F&O BhavCopy file.")
            return

        try:
            self.status_var.set("Starting F&O reconciliation process...")
            self.process_btn.config(state="disabled")
            
            # Load Geneva file
            self.status_var.set("Loading Geneva file...")
            self.geneva_data = self.load_file(self.geneva_file_path, skip_blank_rows=True)
            
            if self.geneva_data is None or len(self.geneva_data) == 0:
                messagebox.showerror("Error", "Geneva file is empty or could not be loaded.")
                self.status_var.set("Geneva file loading failed")
                return
            
            self.status_var.set(f"Geneva file loaded ({len(self.geneva_data)} records)")
            
            # Load BhavCopy file
            self.status_var.set("Loading BhavCopy file...")
            self.bhavcopy_data = self.load_file(self.bhavcopy_file_path, skip_blank_rows=True)
            
            if self.bhavcopy_data is None or len(self.bhavcopy_data) == 0:
                messagebox.showerror("Error", "BhavCopy file is empty or could not be loaded.")
                self.status_var.set("BhavCopy file loading failed")
                return
            
            self.status_var.set(f"BhavCopy file loaded ({len(self.bhavcopy_data)} records)")
            
            # Files will be processed in export_results
            self.status_var.set("Files ready for processing")
            
            # TODO: Implement reconciliation logic
            self.reconciliation_results = {}
            
            self.export_btn.config(state="normal")
            self.status_var.set("Reconciliation completed. Ready to export.")
            
            summary_msg = "Reconciliation completed!\n\n"
            summary_msg += f"Geneva: {len(self.geneva_data)} records\n"
            summary_msg += f"BhavCopy: {len(self.bhavcopy_data)} records\n"
            summary_msg += f"Holdings: {len(self.holdings_files)} files selected"
            
            messagebox.showinfo("Success", summary_msg)
            
        except Exception as e:
            messagebox.showerror("Error", f"Reconciliation failed:\n{str(e)}")
            self.status_var.set("Reconciliation failed")
        finally:
            self.process_btn.config(state="normal")

    def load_file(self, file_path, skip_blank_rows=True):
        """Load a single file and return DataFrame"""
        try:
            data = read_file(file_path, skip_blank_rows=skip_blank_rows)
            return data
        except Exception as e:
            raise Exception(f"Failed to load file {os.path.basename(file_path)}: {str(e)}")


    def format_excel_sheet(self, worksheet, num_rows, num_cols):
        """Apply formatting to Excel worksheet"""
        # Header formatting
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        no_border = Border()
        
        # Data formatting - Light blue zebra
        light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Apply header formatting (row 1)
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = no_border
        
        # Apply zebra striping to data rows
        for row in range(2, num_rows + 2):
            if row % 2 == 0:
                for col in range(1, num_cols + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = light_blue_fill

    def export_results(self):
        """Export F&O reconciliation results with robust error handling"""
        # Ask for export path first
        export_path = self._get_export_path()
        if not export_path:
            return
        
        # Create and show loading spinner
        loader = LoadingSpinner(self, text="Exporting results...")
        
        def task():
            """Run the heavy export work in a background thread"""
            try:
                # Load configuration and initialize data structure
                fund_map = self._load_fund_mapping()
                data_dict = self._initialize_data_dict()
                
                
                # Process all files
                holdings_data = self._load_holdings_files()
                if not holdings_data:
                    loader.close()
                    messagebox.showwarning("No Files", "No holding files were successfully loaded.")
                    return
                
                # Add BhavCopy file FIRST (so price dict is available for Geneva processing)
                if self.bhavcopy_file_path and os.path.exists(self.bhavcopy_file_path):
                    try:
                        bhavcopy_df = read_file(self.bhavcopy_file_path, skip_blank_rows=True)
                        bhavcopy_item = {
                            "filename": os.path.basename(self.bhavcopy_file_path),
                            "filepath": self.bhavcopy_file_path,
                            "data": bhavcopy_df,
                            "type": "BHAVCOPY"
                        }
                        holdings_data.append(bhavcopy_item)
                    except Exception as e:
                        loader.close()
                        messagebox.showerror("BhavCopy File Error", f"Failed to load BhavCopy file:\n{str(e)}")
                        return
                
                # Add Geneva file as an item if it exists
                if self.geneva_file_path and os.path.exists(self.geneva_file_path):
                    try:
                        geneva_df = read_file(self.geneva_file_path, skip_blank_rows=True)
                        geneva_item = {
                            "filename": os.path.basename(self.geneva_file_path),
                            "filepath": self.geneva_file_path,
                            "data": geneva_df,
                            "type": "GENEVA"
                        }
                        holdings_data.append(geneva_item)
                    except Exception as e:
                        loader.close()
                        messagebox.showerror("Geneva File Error", f"Failed to load Geneva file:\n{str(e)}")
                        return
                
                # Process each file (including Geneva and BhavCopy)
                for item in holdings_data:
                    print('item:', item['filename'])
                    self._process_holdings_file(item, fund_map, data_dict)
                
                # breakpoint()
                # Export to Excel
                zip_path = self._export_to_excel(export_path, data_dict)
                
                # Close spinner first
                loader.close()
                
                # Then show success
                messagebox.showinfo(
                    "Success",
                    f"Results exported to ZIP:\n{os.path.basename(zip_path)}\n\nExport completed successfully."
                )
                
            except Exception as e:
                raise e
                loader.close()
                messagebox.showerror("Export Error", f"Export failed:\n{str(e)}")
                self.status_var.set("Export failed")
        
        # Run the task in a background thread
        threading.Thread(target=task, daemon=True).start()

    def _get_export_path(self):
        """Get export file path from user"""
        return filedialog.asksaveasfilename(
            title="Export F&O Reconciliation Results",
            defaultextension=".zip",
            filetypes=[("ZIP Files", "*.zip"), ("All Files", "*.*")],
            initialfile=f"FO_Reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        )

    def _load_fund_mapping(self):
        """Load fund mapping from consolidated data"""
        try:
            from my_app.file_utils import get_app_directory
            app_dir = get_app_directory()
            consolidated_path = os.path.join(app_dir, "consolidated_data.json")
            
            if os.path.exists(consolidated_path):
                with open(consolidated_path, "r") as f:
                    fund_filename_map = json.load(f)
                return fund_filename_map.get("fund_filename_map", {})
            else:
                return {}
        except Exception as e:
            return {}

    def _load_geneva_custodian_mapping(self):
        """Load Geneva custodian mapping from consolidated data"""
        try:
            from my_app.file_utils import get_app_directory
            app_dir = get_app_directory()
            consolidated_path = os.path.join(app_dir, "consolidated_data.json")
            
            if os.path.exists(consolidated_path):
                with open(consolidated_path, "r") as f:
                    consolidated_data = json.load(f)
                return consolidated_data.get("geneva_custodian_mapping", {})
            else:
                return {}
        except Exception as e:
            return {}

    def _initialize_data_dict(self):
        """Initialize data dictionary for different holding types"""
        return {
            "CDS_HOLDINGS": [],
            "REGULAR_HOLDINGS": [],
            "GENEVA_HOLDINGS": [],
            "NSE_F_AND_O_BHAVCOPY": []
        }

    def _load_holdings_files(self):
        """Load all holdings files with error handling"""
        holdings_data = []
        
        for file_path in self.holdings_files:
            try:
                file_name = os.path.basename(file_path)
                self.status_var.set(f"Loading {file_name}...")
                
                data = self.load_file(file_path, skip_blank_rows=True)
                holdings_data.append({
                    "filename": file_name,
                    "data": data,
                    "file_path": file_path
                })
                
            except Exception as e:
                messagebox.showwarning("File Error", f"Failed to load {os.path.basename(file_path)}:\n{str(e)}")
                
        return holdings_data

    def _get_fund_name(self, filename, fund_map):
        """Get fund name for the given filename"""
        for fund_code, details in fund_map.items():
            if fund_code in filename:
                fund_names = details.get("Fund Names", {})
                if "CDS" in filename.upper() and "CDS" in fund_names:
                    return fund_names["CDS"]
                else:
                    return fund_names.get("Default", fund_code)
        return ""

    def _convert_to_numeric(self, value):
        """Convert value to appropriate numeric type for Excel"""
        if pd.isna(value) or value == '' or value is None:
            return None
        elif isinstance(value, (int, float)):
            return float(value)
        else:
            try:
                cleaned = str(value).replace(',', '').replace(' ', '').strip()
                if cleaned in ['', '-', 'N/A']:
                    return None
                return float(cleaned)
            except (ValueError, TypeError, decimal.InvalidOperation):
                return str(value)

    def _create_concatenate_code(self, row):
        """Create concatenate code from row data"""
        try:
            exchange = str(row['Exchange']).strip()
            underlying_code = str(row['UnderlyingCode']).strip()
            expiry_date = parse_expiry_date(row['ExpiryDate'])

            option_type = str(row['OptionType']).strip()[0]
            strike_price = int(row['StrikePrice'])
            concatenate_code = f"{exchange}{underlying_code}{expiry_date}{option_type}{strike_price}"
            return concatenate_code
        except Exception as e:
            return ""

    def _create_bhavcopy_concatenated_key(self, row):
        """Create concatenated key for BhavCopy data using NSE+SYMBOL+EXPIRY_DT+OPTION_TYP+STRIKE_PR format"""
        try:
            # Format: NSE + SYMBOL + EXPIRY_DT (yyyymmdd) + OPTION_TYP (first char) + STRIKE_PR
            symbol = str(row.get('SYMBOL', '')).strip()
            expiry_dt = pd.to_datetime(row.get('EXPIRY_DT')).strftime('%Y%m%d')
            option_typ = str(row.get('OPTION_TYP', '')).strip()
            option_typ_first_char = option_typ[0] if option_typ else ''
            strike_pr = str(int(row.get('STRIKE_PR', 0)))
            
            return f"NSE{symbol}{expiry_dt}{option_typ_first_char}{strike_pr}"
        except Exception as e:
            return ""

    def _calculate_net_difference(self, row):
        """Calculate net buy - net sell with Decimal precision"""
        try:
            netbuy_decimal = Decimal(str(row['NetBuy'])) if pd.notna(row['NetBuy']) else Decimal('0')
            netsell_decimal = Decimal(str(row['NetSell'])) if pd.notna(row['NetSell']) else Decimal('0')
            return float(netbuy_decimal - netsell_decimal)
        except Exception as e:
            return 0.0

    def _process_holdings_file(self, item, fund_map, data_dict):
        """Process a single holdings file and populate data_dict"""
        filename = item["filename"]
        df = item["data"]
        file_type = item.get("type", "HOLDINGS")  # Default to HOLDINGS if not specified
        
        # Handle Geneva and BhavCopy files differently
        if file_type == "GENEVA":
            self._process_geneva_data(df, data_dict)
            return
        elif file_type == "BHAVCOPY":
            self._process_bhavcopy_data(df, data_dict)
            return
        
        # Process regular holdings files
        fund_name = self._get_fund_name(filename, fund_map)
        print(f"Processing {filename} as {fund_name}")

        for _, row in df.iterrows():
            try:
                # Build row values with proper numeric conversion
                row_values = [fund_name]
                for v in row:
                    row_values.append(self._convert_to_numeric(v))

                # Add calculated fields
                concatenate_code = self._create_concatenate_code(row)
                client_name = str(row['ClientName']).strip()
                unique_code = f"{concatenate_code}{client_name}"
                net_difference = self._calculate_net_difference(row)
                
                row_values.extend([concatenate_code, "", unique_code, net_difference])
                
                # Categorize data based on filename
                self._categorize_data(filename, row_values, data_dict)

            except Exception as e:
                continue
        # breakpoint()

    def _process_geneva_data(self, df, data_dict):
        """Process Geneva data and add to data_dict"""
        try:
            self.status_var.set("Processing Geneva data...")
            
            
            # Load Geneva custodian mapping from consolidated_data.json
            geneva_mapping = self._load_geneva_custodian_mapping()
            
            # Ensure bhavcopy_price_dict exists (should be created by BhavCopy processing)
            if not hasattr(self, 'bhavcopy_price_dict') or not self.bhavcopy_price_dict:
                self.bhavcopy_price_dict = {}
            
            # Create dynamic headers that match the actual data structure
            # Use the actual DataFrame columns + the 4 extra columns we're adding
            self.dynamic_geneva_headers = list(df.columns) + ['UniqueCode', 'Fund Name IN Custody Holding', 'Nse Price', 'Diff']
            
            for _, row in df.iterrows():
                row_values = []
                for v in row:
                    row_values.append(self._convert_to_numeric(v))
                
                
                # Get portfolio value and map it using geneva_custodian_mapping
                portfolio_value = str(row['Portfolio']).strip()
                fund_name_in_custody_holding = geneva_mapping.get(portfolio_value, portfolio_value)
                
                # Add additional calculated fields for Geneva
                investment_desc = str(row['Investment Description']).strip()
                unique_code = f"{investment_desc}{fund_name_in_custody_holding}"
                investment = str(row['Investment']).strip()
                nse_price = self.bhavcopy_price_dict.get(investment)
                # Handle None values for nse_price
                if nse_price is None:
                    nse_price = 0.0
                
                market_local_price = row['Market Price Local']
                # Handle None values for market_local_price
                if pd.isna(market_local_price) or market_local_price is None:
                    market_local_price = 0.0
                
                diff = float(market_local_price) - float(nse_price)
                row_values.extend([unique_code, fund_name_in_custody_holding, nse_price, diff])
                
                
                # Add to Geneva holdings
                data_dict["GENEVA_HOLDINGS"].append(row_values)
                
            
        except Exception as e:
            messagebox.showerror("Geneva Processing Error", f"Failed to process Geneva data:\n{str(e)}")
            self.status_var.set("Geneva data processing failed")

    def _process_bhavcopy_data(self, df, data_dict):
        """Process BhavCopy data and add to data_dict"""
        try:
            self.status_var.set("Processing BhavCopy data...")
            
            # Initialize dictionary for concatenated key and closing price
            self.bhavcopy_price_dict = {}
            
            for _, row in df.iterrows():
                row_values = []
                for v in row:
                    row_values.append(self._convert_to_numeric(v))
                
                # Create concatenated key and store closing price
                concatenated_key = self._create_bhavcopy_concatenated_key(row)
                closing_price = self._convert_to_numeric(row.get('CLOSE', 0))
                
                self.bhavcopy_price_dict[concatenated_key] = closing_price

                row_values.extend([concatenated_key])
                
                # Add to NSE F&O BhavCopy
                data_dict["NSE_F_AND_O_BHAVCOPY"].append(row_values)
                
            
        except Exception as e:
            messagebox.showerror("BhavCopy Processing Error", f"Failed to process BhavCopy data:\n{str(e)}")
            self.status_var.set("BhavCopy data processing failed")

    def _categorize_data(self, filename, row_values, data_dict):
        """Categorize row data based on filename"""
        fname_upper = filename.upper()
        
        if "CDS" in fname_upper:
            data_dict["CDS_HOLDINGS"].append(row_values)
        else:
            data_dict["REGULAR_HOLDINGS"].append(row_values)
        # elif "GENEVA" in fname_upper:
        #     data_dict["GENEVA_HOLDINGS"].append(row_values)
        # elif "NSE" in fname_upper or "BHAVCOPY" in fname_upper:
        #     data_dict["NSE_F_AND_O_BHAVCOPY"].append(row_values)

    def _export_to_excel(self, export_path, data_dict):
        """Export processed data to Excel file"""
        self.status_var.set("Exporting results...")
        
        # Check if any data exists
        total_records = sum(len(data_dict[key]) for key in data_dict)
        if total_records == 0:
            messagebox.showwarning("No Data", "No data found to export. Please check your files.")
            return
        
        # Prepare sheets to create
        sheets_to_create = self._prepare_excel_sheets(data_dict)
        
        # --- Create in-memory Excel files ---
        processed_buffer = io.BytesIO()
        cds_reg_buffer = io.BytesIO()
        sheets_created = 0
        zip_path = ""

        try:
            # 1️⃣ Write processed data to first Excel
            with pd.ExcelWriter(processed_buffer, engine='openpyxl') as writer:
                for sheet_name, df in sheets_to_create:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    self.format_excel_sheet(writer.sheets[sheet_name], len(df), len(df.columns))
                    sheets_created += 1
            
            # 2️⃣ Write CDS and REG Geneva data to second Excel
            with pd.ExcelWriter(cds_reg_buffer, engine='openpyxl') as writer:
                cds_df = pd.DataFrame(self.cds_geneva_data, columns=CDS_GENEVA_HEADER_LIST)
                cds_df.to_excel(writer, sheet_name="CDS_Geneva_Data", index=False)
                self.format_excel_sheet(writer.sheets["CDS_Geneva_Data"], len(cds_df), len(cds_df.columns))

                reg_df = pd.DataFrame(self.reg_geneva_data, columns=REG_GENEVA_HEADER_LIST)
                reg_df.to_excel(writer, sheet_name="REG_Geneva_Data", index=False)
                self.format_excel_sheet(writer.sheets["REG_Geneva_Data"], len(reg_df), len(reg_df.columns))

            # --- Create ZIP archive ---
            # Ensure path ends with .zip
            if not export_path.lower().endswith('.zip'):
                zip_path = os.path.splitext(export_path)[0] + ".zip"
            else:
                zip_path = export_path
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
                zipf.writestr("Processed_Data.xlsx", processed_buffer.getvalue())
                zipf.writestr("Geneva_Data.xlsx", cds_reg_buffer.getvalue())
        except Exception as e:
            raise e

        self.status_var.set(f"Results exported to {os.path.basename(zip_path)}")
        
        return zip_path

    def geneva_with_cds_and_reg(self, cds_row=None, regular_row=None, geneva_row=None,type=None):
        """Combine Geneva with either CDS or Regular rows depending on input."""

        # --- Common Geneva fields ---
        geneva_fields = [
            "Portfolio", "Investment", "Investment Description", "Settled Quantity",
            "Unit Cost", "Currency", "Cost Local", "Market Price Local",
            "Cost Book", "Market Value Local", "UniqueCode", "Market Value Book",
            "Unrealized G/L Book", "Accrued Interest Local", "Cust Account",
            "Strategy", "Fund Name IN Custody Holding", "Nse Price", "Diff"
        ]

        # --- CDS fields ---
        cds_fields = [
            "Source.Name", "Date", "QuantityUnit", "Exchange", "ClientCode",
            "TradingCode", "UnderlyingCode", "ClientName", "UnderlyingName",
            "InstrumentType", "ExpiryDate", "OptionType", "StrikePrice",
            "OpenBuy", "OpenSell", "TradedBuy", "DayBuyValue", "TradedSell",
            "DaySellValue", "ExcerciseQty", "AllocationQty", "NetBuy", "NetSell",
            "ContractSettlementPrice", "Settlement Price", "BloombergCodes",
            "Concatenate", "Fund Name In Geneva", "UniqueCode", "Netbuy-Netsell"
        ]

        # --- Regular fields ---
        regular_fields = [
            "Source.Name", "Date", "Exchange", "ClientName", "ClientCode",
            "OptionType", "UnderlyingCode", "UnderlyingName", "InstrumentType",
            "StrikePrice", "ExpiryDate", "OpenBuy", "OpenSell", "TradedSell",
            "TradedBuy", "DayBuyValue", "DaySellValue", "ExcerciseQty",
            "AllocationQty", "NetBuy", "NetSell", "ContractSettlementPrice",
            "ClosingPrice", "BloombergCD", "Concatenate", "Fund Name In Geneva",
            "UniqueCode", "NerBuy-Netsell"
        ]

        # Helper to extract field values safely
        def extract_fields(row, fields):
            return [row.get(f, "") for f in fields] if row else [""] * len(fields)

        # Build Geneva portion
        g_row = extract_fields(geneva_row, geneva_fields)

        # Determine if CDS or Regular and build accordingly
        if cds_row:
            merged = g_row + extract_fields(cds_row, cds_fields)
        elif regular_row:
            merged = g_row + extract_fields(regular_row, regular_fields)
        elif type=="REGULAR":
            merged = g_row + [""] * (len(regular_fields))  # Default to Regular fields length
        elif type=="CDS":
            merged = g_row + [""] * (len(cds_fields))  # Default to CDS fields length
        
        return merged

    def _prepare_excel_sheets(self, data_dict):
        """Prepare list of sheets to create in Excel"""
        sheets_to_create = []
        
        # Define sheet configurations
        sheet_configs = [
            ("CDS_HOLDINGS", CDS_HOLDINGS_HEADER),
            ("REGULAR_HOLDINGS", REGULAR_HOLDINGS_HEADERS),
            ("NSE_F_AND_O_BHAVCOPY", NSE_F_AND_O_BHAVCOPY)
        ]
        
        # Handle Geneva separately with dynamic headers
        if data_dict["GENEVA_HOLDINGS"]:
            # Use the first row to determine the number of columns
            first_row = data_dict["GENEVA_HOLDINGS"][0]
            num_columns = len(first_row)
            
            # Create dynamic headers for Geneva
            if hasattr(self, 'dynamic_geneva_headers') and len(self.dynamic_geneva_headers) == num_columns:
                geneva_headers = self.dynamic_geneva_headers
            else:
                # Fallback: create generic headers
                geneva_headers = [f"Column_{i+1}" for i in range(num_columns)]
            
            df = pd.DataFrame(data_dict["GENEVA_HOLDINGS"], columns=geneva_headers)
            sheets_to_create.append(("GENEVA_HOLDINGS", df))
        
        # Process other sheets
        for key, headers in sheet_configs:
            if data_dict[key]:
                print(f"Preparing sheet: {key}")
                print(f"data: {len(data_dict[key][0])} {len(headers)}")
                df = pd.DataFrame(data_dict[key], columns=headers)
                sheets_to_create.append((key, df))

        # Add Master File sheet last
        master_file_data = self._create_master_file_data()
        if master_file_data:
            master_df = pd.DataFrame(master_file_data[1:], columns=master_file_data[0])
            sheets_to_create.append(("MASTER_FILE", master_df))
        
        # Fallback: create summary sheet if no data
        if not sheets_to_create:
            summary_df = pd.DataFrame({'Message': ['No data found to export']})
            sheets_to_create.append(('Summary', summary_df))
        
        # Prepare additional raw data sheets
        print("Preparing additional raw data sheets...")
        self._build_geneva_reconciliation_data(data_dict)
        
        return sheets_to_create

    def _build_geneva_reconciliation_data(self, data_dict):
        """Build CDS and REGULAR reconciliation with Geneva data"""
        # Convert rows to dictionaries only once
        cds_rows = [dict(zip(CDS_HOLDINGS_HEADER, row)) for row in data_dict["CDS_HOLDINGS"]]
        regular_rows = [dict(zip(REGULAR_HOLDINGS_HEADERS, row)) for row in data_dict["REGULAR_HOLDINGS"]]
        geneva_rows = [dict(zip(self.dynamic_geneva_headers, row)) for row in data_dict["GENEVA_HOLDINGS"]]
        
        # Load mappings
        geneva_mapping = self._load_geneva_custodian_mapping()
        
        # Build Geneva investment-to-portfolios mapping
        _geneva = defaultdict(list)
        for row in geneva_rows:
            inv = row.get("Investment")
            port = row.get("Portfolio")
            if inv and port and port not in _geneva[inv]:
                _geneva[inv].append(port)
        
        # Build Geneva lookup
        geneva_lookup = {f"{r.get('Investment')}_{r.get('Portfolio')}": r for r in geneva_rows}
        
        # Helper function to map client name to portfolio
        def map_client_to_portfolio(client_name, concatenate):
            """Maps client name to portfolio using geneva_mapping"""
            match_keys = [k for k, v in geneva_mapping.items() if v == client_name]
            portfolios = _geneva.get(concatenate, [])
            
            if match_keys:
                # Prefer keys that are actually present in portfolios
                preferred = [k for k in match_keys if k in portfolios]
                return preferred[0] if preferred else match_keys[0]
            elif len(portfolios) == 1:
                return portfolios[0]
            return None
        
        # Build CDS lookup
        cds_lookup = {}
        for r in cds_rows:
            mapped_key = map_client_to_portfolio(r.get("ClientName"), r.get("Concatenate"))
            lookup_key = f"{r.get('Concatenate')}_{mapped_key}"
            cds_lookup[lookup_key] = r
        
        # Build REGULAR lookup
        regular_lookup = {}
        for r in regular_rows:
            mapped_key = map_client_to_portfolio(r.get("ClientName"), r.get("Concatenate"))
            lookup_key = f"{r.get('Concatenate')}_{mapped_key}"
            regular_lookup[lookup_key] = r
        
        # Merge CDS + Geneva
        cds_geneva_keys = set(cds_lookup.keys()) | set(geneva_lookup.keys())
        cds_geneva_data = []
        
        for key in cds_geneva_keys:
            cds_row = cds_lookup.get(key)
            geneva_row = geneva_lookup.get(key)
            
            # Calculate values safely
            cds_val = float(cds_row.get('Netbuy-Netsell', 0) or 0) if cds_row else 0
            geneva_val = float(geneva_row.get('Traded Quantity', 0) or 0) if geneva_row else 0
            diff = float(geneva_val - cds_val)
            
            investment = str(geneva_row.get('Investment') or "") if geneva_row else ""
            
            # Include if both exist, or Geneva is NSECR, or only CDS exists
            if (geneva_row and cds_row) or (geneva_row and "NSECR" in investment) or cds_row:
                _row = self.geneva_with_cds_and_reg(
                    cds_row=cds_row,
                    regular_row=None,
                    geneva_row=geneva_row,
                    type="CDS"
                )
                cds_geneva_data.append(_row + [geneva_val, diff])
        
        # Merge REGULAR + Geneva (exclude keys already processed in CDS)
        reg_geneva_keys = (set(regular_lookup.keys()) | set(geneva_lookup.keys())) - set(cds_lookup.keys())
        reg_geneva_data = []
        
        for key in reg_geneva_keys:
            reg_row = regular_lookup.get(key)
            geneva_row = geneva_lookup.get(key)
            
            reg_val = float(reg_row.get('NerBuy-Netsell', 0) or 0) if reg_row else 0
            geneva_val = float(geneva_row.get('Traded Quantity', 0) or 0) if geneva_row else 0
            diff = float(geneva_val - reg_val)
            
            investment = str(geneva_row.get('Investment') or "") if geneva_row else ""
            
            # Include if REGULAR exists, or Geneva is NSE but not NSECR
            if reg_row or ("NSE" in investment and "NSECR" not in investment):
                _row = self.geneva_with_cds_and_reg(
                    cds_row=None,
                    regular_row=reg_row,
                    geneva_row=geneva_row,
                    type="REGULAR"
                )
                reg_geneva_data.append(_row + [geneva_val, diff])
        
        # Store results
        self.cds_geneva_data = cds_geneva_data
        self.reg_geneva_data = reg_geneva_data

    def _create_master_file_data(self):
        """Create master file data with Geneva custodian mapping"""
        try:
            # Load Geneva custodian mapping
            geneva_mapping = self._load_geneva_custodian_mapping()
            
            master_data = []
            
            # Add Geneva Custodian Mapping
            master_data.append(["Fund Name in Geneva", "Fund Name in Cust Holding"])
            
            for fund_name_geneva, fund_name_cust_holding in geneva_mapping.items():
                master_data.append([fund_name_geneva, fund_name_cust_holding])
            
            return master_data
            
        except Exception as e:
            messagebox.showerror("Master File Error", f"Failed to create master file data:\n{str(e)}")
            return []