import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import json
from datetime import datetime
import pandas as pd
from my_app.pages.helper import read_file
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from my_app.CONSTANTS import FORMAT_1, FORMAT_2, BHAVCOPY

class ASIOReconciliationPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg="#ecf0f1")

        # Title
        title = tk.Label(self, text="🔄 ASIO Reconciliation", font=("Arial", 20, "bold"), bg="#ecf0f1", fg="#2c3e50")
        title.pack(pady=10)

        # Main content frame
        content_frame = tk.Frame(self, bg="#ecf0f1")
        content_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Left panel - File Browser
        left_panel = tk.Frame(content_frame, bg="#ecf0f1")
        left_panel.pack(side="left", fill="both", expand=True, padx=(0, 10))

        # File Browser Section
        file_browser_frame = tk.LabelFrame(left_panel, text="📁 File Browser", font=("Arial", 12, "bold"), 
                                         bg="#ecf0f1", fg="#2c3e50", padx=10, pady=10)
        file_browser_frame.pack(fill="both", expand=True)

        # Browse buttons frame
        browse_buttons_frame = tk.Frame(file_browser_frame, bg="#ecf0f1")
        browse_buttons_frame.pack(fill="x", pady=(0, 10))

        # Browse Files button
        self.browse_files_btn = tk.Button(browse_buttons_frame, text="📂 Browse Files", 
                                        bg="#3498db", fg="white", font=("Arial", 11, "bold"), 
                                        relief="flat", padx=15, pady=8, command=self.browse_files)
        self.browse_files_btn.pack(side="left", padx=(0, 10))

        # Clear Files button
        self.clear_files_btn = tk.Button(browse_buttons_frame, text="🗑️ Clear All", 
                                       bg="#e74c3c", fg="white", font=("Arial", 11, "bold"), 
                                       relief="flat", padx=15, pady=8, command=self.clear_files)
        self.clear_files_btn.pack(side="left")

        # Files listbox with scrollbar
        listbox_frame = tk.Frame(file_browser_frame, bg="#ecf0f1")
        listbox_frame.pack(fill="both", expand=True)

        # Listbox for displaying selected files
        self.files_listbox = tk.Listbox(listbox_frame, font=("Arial", 10), bg="white", 
                                       selectmode=tk.EXTENDED, relief="solid", bd=1)
        files_scrollbar = tk.Scrollbar(listbox_frame, orient="vertical", command=self.files_listbox.yview)
        self.files_listbox.configure(yscrollcommand=files_scrollbar.set)

        self.files_listbox.pack(side="left", fill="both", expand=True)
        files_scrollbar.pack(side="right", fill="y")

        # File info frame
        file_info_frame = tk.Frame(file_browser_frame, bg="#ecf0f1")
        file_info_frame.pack(fill="x", pady=(10, 0))

        self.file_count_label = tk.Label(file_info_frame, text="No files selected", 
                                       font=("Arial", 10), bg="#ecf0f1", fg="#7f8c8d")
        self.file_count_label.pack(side="left")

        # Right panel - Geneva File Browser
        right_panel = tk.Frame(content_frame, bg="#ecf0f1")
        right_panel.pack(side="right", fill="both", expand=True)

        # Geneva File Browser Section
        geneva_frame = tk.LabelFrame(right_panel, text="🏦 Geneva File Browser", font=("Arial", 12, "bold"), 
                                   bg="#ecf0f1", fg="#2c3e50", padx=10, pady=10)
        geneva_frame.pack(fill="both", expand=True)

        # Geneva file path frame
        geneva_path_frame = tk.Frame(geneva_frame, bg="#ecf0f1")
        geneva_path_frame.pack(fill="x", pady=(0, 10))

        tk.Label(geneva_path_frame, text="Geneva File:", font=("Arial", 11), bg="#ecf0f1", fg="#2c3e50").pack(anchor="w")
        
        geneva_input_frame = tk.Frame(geneva_path_frame, bg="#ecf0f1")
        geneva_input_frame.pack(fill="x", pady=(5, 0))

        self.geneva_path_var = tk.StringVar()
        self.geneva_path_entry = tk.Entry(geneva_input_frame, textvariable=self.geneva_path_var, 
                                        font=("Arial", 10), width=50, relief="solid", bd=1)
        self.geneva_path_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        self.browse_geneva_btn = tk.Button(geneva_input_frame, text="Browse", 
                                         bg="#9b59b6", fg="white", font=("Arial", 10, "bold"), 
                                         relief="flat", padx=12, pady=4, command=self.browse_geneva_file)
        self.browse_geneva_btn.pack(side="right")

        # Geneva file info
        self.geneva_info_label = tk.Label(geneva_frame, text="No Geneva file selected", 
                                        font=("Arial", 10), bg="#ecf0f1", fg="#7f8c8d")
        self.geneva_info_label.pack(anchor="w", pady=(10, 0))

        # Process buttons frame
        process_frame = tk.Frame(self, bg="#ecf0f1")
        process_frame.pack(fill="x", padx=20, pady=20)

        # Process button
        self.process_btn = tk.Button(process_frame, text="🔄 Start Reconciliation", 
                                   bg="#27ae60", fg="white", font=("Arial", 12, "bold"), 
                                   relief="flat", padx=20, pady=10, command=self.start_reconciliation)
        self.process_btn.pack(side="left", padx=(0, 10))

        # Export button
        self.export_btn = tk.Button(process_frame, text="📊 Export Results", 
                                  bg="#8e44ad", fg="white", font=("Arial", 12, "bold"), 
                                  relief="flat", padx=20, pady=10, command=self.export_results,
                                  state="disabled")
        self.export_btn.pack(side="left")

        # Status bar
        self.status_var = tk.StringVar(value="Ready - Select files and Geneva file to begin reconciliation")
        status_label = tk.Label(self, textvariable=self.status_var, font=("Arial", 10), 
                              bg="#ecf0f1", fg="#7f8c8d", anchor="w")
        status_label.pack(fill="x", padx=20, pady=(0, 10))

        # Data storage
        self.selected_files = []
        self.geneva_file_path = ""
        self.reconciliation_results = None
        self.geneva_data = None
        self.holding_files_data = []

    def browse_files(self):
        """Browse and select multiple files"""
        file_types = [
            ("All Supported Files", "*.xlsx *.xls *.csv"),
            ("Excel Files", "*.xlsx *.xls"),
            ("CSV Files", "*.csv"),
            ("All Files", "*.*")
        ]
        
        files = filedialog.askopenfilenames(
            title="Select Files for Reconciliation",
            filetypes=file_types
        )
        
        if files:
            self.selected_files = list(files)
            self.update_files_listbox()
            self.update_file_count()
            self.status_var.set(f"Selected {len(self.selected_files)} files")

    def browse_geneva_file(self):
        """Browse for Geneva file"""
        file_types = [
            ("All Supported Files", "*.xlsx *.xls *.csv"),
            ("Excel Files", "*.xlsx *.xls"),
            ("CSV Files", "*.csv"),
            ("All Files", "*.*")
        ]
        
        file_path = filedialog.askopenfilename(
            title="Select Geneva File",
            filetypes=file_types
        )
        
        if file_path:
            self.geneva_file_path = file_path
            self.geneva_path_var.set(file_path)
            file_name = os.path.basename(file_path)
            self.geneva_info_label.config(
                text=f"Selected: {file_name}",
                fg="#27ae60"
            )
            self.status_var.set(f"Geneva file selected: {file_name}")

    def clear_files(self):
        """Clear all selected files"""
        self.selected_files = []
        self.holding_files_data = []
        self.update_files_listbox()
        self.update_file_count()
        self.status_var.set("Files cleared")

    def update_files_listbox(self):
        """Update the files listbox with selected files"""
        self.files_listbox.delete(0, tk.END)
        for file_path in self.selected_files:
            file_name = os.path.basename(file_path)
            self.files_listbox.insert(tk.END, file_name)

    def update_file_count(self):
        """Update the file count label"""
        count = len(self.selected_files)
        if count == 0:
            self.file_count_label.config(text="No files selected", fg="#7f8c8d")
        else:
            self.file_count_label.config(text=f"{count} file(s) selected", fg="#27ae60")

    def start_reconciliation(self):
        """Start the reconciliation process"""
        if not self.selected_files:
            messagebox.showwarning("No Files", "Please select files to reconcile.")
            return
        
        if not self.geneva_file_path:
            messagebox.showwarning("No Geneva File", "Please select a Geneva file.")
            return

        try:
            self.status_var.set("Starting reconciliation process...")
            self.process_btn.config(state="disabled")
            
            # Load Geneva file
            self.status_var.set("Loading Geneva file...")
            self.geneva_data = self.load_geneva_file()
            
            if self.geneva_data is None or len(self.geneva_data) == 0:
                messagebox.showerror("Error", "Geneva file is empty or could not be loaded.")
                self.status_var.set("Geneva file loading failed")
                return
            
            self.status_var.set(f"Geneva file loaded ({len(self.geneva_data)} records)")
            
            # Load all holding files
            self.status_var.set("Loading holding files...")
            self.holding_files_data = self.load_holding_files()
            
            if len(self.holding_files_data) == 0:
                messagebox.showwarning("No Files", "No holding files were successfully loaded.")
                self.status_var.set("No files loaded")
                return
            
            self.status_var.set(f"Loaded {len(self.holding_files_data)} holding files")
            
            # TODO: Implement your reconciliation logic here
            # You have access to:
            # - self.geneva_data: DataFrame from Geneva file
            # - self.holding_files_data: List of dicts with 'format' and 'data' (DataFrame)
            
            # Store results for export (format: {file_name: mismatches_dict})
            self.reconciliation_results = {}
            
            # Example structure for results:
            # self.reconciliation_results = {
            #     "Holding File 1.xlsx": {
            #         "ISIN1": {"Geneva": 100, "Holding": 95, "Difference": 5}
            #     }
            # }
            
            self.export_btn.config(state="normal")
            self.status_var.set("Reconciliation completed. Ready to export.")
            messagebox.showinfo("Success", f"Reconciliation completed!\n\nGeneva: {len(self.geneva_data)} records\nHolding files: {len(self.holding_files_data)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Reconciliation failed:\n{str(e)}")
            self.status_var.set("Reconciliation failed")
        finally:
            self.process_btn.config(state="normal")

    def load_header_configs(self):
        """Load header configurations from consolidated_data.json"""
        try:
            from my_app.file_utils import get_app_directory
            app_dir = get_app_directory()
            consolidated_path = os.path.join(app_dir, "consolidated_data.json")
            
            if os.path.exists(consolidated_path):
                with open(consolidated_path, "r") as f:
                    consolidated_data = json.load(f)
                    return {
                        'format_1': consolidated_data.get("asio_format_1_headers", {}),
                        'format_2': consolidated_data.get("asio_format_2_headers", {}),
                        'bhavcopy': consolidated_data.get("asio_bhavcopy_headers", {})
                    }
        except Exception:
            pass
        
        return {'format_1': {}, 'format_2': {}, 'bhavcopy': {}}

    def load_portfolio_mapping(self):
        """Load ASIO portfolio mapping from consolidated data"""
        try:
            from my_app.file_utils import get_app_directory
            app_dir = get_app_directory()
            consolidated_path = os.path.join(app_dir, "consolidated_data.json")
            
            if os.path.exists(consolidated_path):
                with open(consolidated_path, "r") as f:
                    consolidated_data = json.load(f)
                    return consolidated_data.get("asio_portfolio_mapping", {})
        except Exception:
            pass
        
        return {}

    def detect_format(self, df, header_configs):
        """
        Detect which format the DataFrame matches
        
        Priority: Format 1 → Format 2 → BhavCopy
        
        Args:
            df: DataFrame to check
            header_configs: Dict with 'format_1', 'format_2', 'bhavcopy' headers
            
        Returns:
            str: 'Format 1', 'Format 2', 'BhavCopy', or None
        """
        file_columns = set(df.columns)
        
        # Check Format 1 - ALL headers must match
        format_1_headers = set(header_configs['format_1'].values())
        if format_1_headers and format_1_headers.issubset(file_columns):
            return FORMAT_1
        
        # Check Format 2 - ALL headers must match
        format_2_headers = set(header_configs['format_2'].values())
        if format_2_headers and format_2_headers.issubset(file_columns):
            return FORMAT_2
        
        # Check BhavCopy - ALL headers must match
        bhavcopy_headers = set(header_configs['bhavcopy'].values())
        if bhavcopy_headers and bhavcopy_headers.issubset(file_columns):
            return BHAVCOPY
        
        return None

    def load_holding_files(self):
        """Function 2: Load all holding files with automatic format detection"""
        header_configs = self.load_header_configs()
        holding_files_data = []
        failed_files = []
        unknown_format_files = []
        
        for file_path in self.selected_files:
            file_name = os.path.basename(file_path)
            self.status_var.set(f"Loading {file_name}...")
            df = None
            detected_format = None
            
            try:
                file_ext = file_path.lower()
                if file_ext.endswith(('.xls', '.xlsx')):
                    df = pd.read_excel(file_path, sheet_name=0, engine='xlrd' if file_ext.endswith('.xls') else 'openpyxl')
                else:
                    df = pd.read_csv(file_path)
                
                if isinstance(df, dict):
                    df = list(df.values())[0]
                
                detected_format = self.detect_format(df, header_configs)
                
                if not detected_format:
                    if file_ext.endswith(('.xls', '.xlsx')):
                        df = pd.read_excel(file_path, sheet_name=0, header=3, engine='xlrd' if file_ext.endswith('.xls') else 'openpyxl')
                    else:
                        df = pd.read_csv(file_path, header=3)
                    
                    if isinstance(df, dict):
                        df = list(df.values())[0]
                    
                    detected_format = self.detect_format(df, header_configs)
                
                if not detected_format:
                    unknown_format_files.append(file_name)
            
            except Exception as e:
                failed_files.append(f"{file_name}: {str(e)[:50]}")
                continue
            
            holding_files_data.append({'format': detected_format, 'data': df})
        
        if failed_files:
            messagebox.showwarning("File Loading Errors", "Failed to load files:\n" + "\n".join(failed_files))
            
        if unknown_format_files:
            messagebox.showwarning("Unknown Format", "Files with unknown format:\n" + "\n".join(unknown_format_files))
            
        return holding_files_data

    def load_geneva_file(self):
        """
        Function 3: Load Geneva file
        
        Returns:
            DataFrame: Geneva data
        """
        try:
            geneva_data = read_file(self.geneva_file_path)
            
            # If multiple sheets returned, get first sheet
            if isinstance(geneva_data, dict):
                geneva_data = list(geneva_data.values())[0]
            
            return geneva_data
        except Exception as e:
            raise Exception(f"Failed to load Geneva file: {str(e)}")

    def format_excel_sheet(self, worksheet, num_rows, num_cols):
        """
        Apply formatting to Excel worksheet
        - Header: Blue, Accent 1, Darker 25% (no borders)
        - Data: Light blue and white zebra striping
        """
        # Header formatting - Blue, Accent 1, Darker 25%
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        no_border = Border()  # Empty border (no borders)
        
        # Data formatting - Light blue zebra
        light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Apply header formatting (row 1)
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = no_border  # Remove border from header
        
        # Apply zebra striping to data rows (alternating light blue and white)
        for row in range(2, num_rows + 2):  # Start from row 2 (data rows)
            if row % 2 == 0:  # Even rows - light blue
                for col in range(1, num_cols + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = light_blue_fill

    def export_results(self):
        """
        Function 4: Export reconciliation results
        
        """
        # if not self.reconciliation_results:
        #     messagebox.showwarning("No Results", "No reconciliation results to export.")
        #     return

        try:
            # Ask for export location
            export_path = filedialog.asksaveasfilename(
                title="Export Reconciliation Results",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile=f"ASIO_Reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if export_path:
                self.status_var.set("Exporting results...")
                
                # Access Geneva data
                geneva_data = self.geneva_data
                
                # Access holding files data
                holding_files_data = self.holding_files_data
                
                # Access holding files
                format_1_results = []  # Collect all Format 1 results
                format_2_results = []  # Collect all Format 2 results
                bhavcopy_results = []  # Collect all BhavCopy results
                
                for holding in self.holding_files_data:
                    format_type = holding['format']  # 'Format 1', 'Format 2', 'BhavCopy'
                    holding_df = holding['data']      # DataFrame
                    
                    # Your reconciliation logic here
                    if format_type == FORMAT_1:
                        # Handle Format 1 reconciliation with Geneva - Manual implementation
                        
                        # Load portfolio mapping (Portfolio Code -> Holding Name)
                        portfolio_mapping = self.load_portfolio_mapping()
                        
                        # Create Geneva dict: Cln Name-ISIN -> Traded Quantity
                        geneva_dict = {}
                        for _, geneva_row in geneva_data.iterrows():
                            portfolio_code = str(geneva_row['Portfolio']).strip()
                            investment = str(geneva_row['Investment']).strip()
                            
                            # Map Portfolio Code to Holding Name (Cln Name)
                            cln_name = portfolio_mapping.get(portfolio_code, portfolio_code)
                            
                            # Create key: Cln Name-ISIN
                            key = f"{cln_name}-{investment}"
                            geneva_dict[key] = geneva_row['Traded Quantity']
                        
                        # Match and reconcile manually
                        reconciled_records = []
                        for _, holding_row in holding_df.iterrows():
                            cln_name = str(holding_row['Cln Name']).strip()
                            instr_isin = str(holding_row['Instr ISIN']).strip()
                            
                            # Create key: Cln Name-ISIN
                            key = f"{cln_name}-{instr_isin}"
                            
                            # Check if this key exists in Geneva
                            if key in geneva_dict:
                                traded_quantity = geneva_dict[key]
                                saleable = holding_row['Saleable']
                                
                                # Calculate difference: Traded Quantity - Saleable
                                difference = traded_quantity - saleable
                                
                                # Create reconciled record
                                record = {
                                    'Client Name': cln_name,
                                    'ISIN': instr_isin,
                                    'Settled Position': holding_row['Settled Position'],
                                    'Pending Purchase': holding_row['Pending Purchase'],
                                    'Pending Sale': holding_row['Pending Sale'],
                                    'Pending CA Entitlements': holding_row['Pending CA Entitlements'],
                                    'Market Price Date': holding_row['Market Price Date'],
                                    'Market Price': holding_row['Market Price'],
                                    'Saleable': saleable,
                                    'Geneva Quantity': traded_quantity,
                                    'Difference': difference
                                }
                                
                                reconciled_records.append(record)
                        
                        # Add to Format 1 results
                        if reconciled_records:
                            format_1_results.extend(reconciled_records)
                    elif format_type == FORMAT_2:
                        # Handle Format 2 reconciliation with Geneva - Manual implementation
                        
                        # Load portfolio mapping (Portfolio Code -> Holding Name)
                        portfolio_mapping = self.load_portfolio_mapping()
                        
                        # Create Geneva dict: Securities account name-ISIN -> Traded Quantity
                        geneva_dict = {}
                        for _, geneva_row in geneva_data.iterrows():
                            portfolio_code = str(geneva_row['Portfolio']).strip()
                            investment = str(geneva_row['Investment']).strip()
                            
                            # Map Portfolio Code to Holding Name (Securities account name)
                            securities_account_name = portfolio_mapping.get(portfolio_code, portfolio_code)
                            
                            # Create key: Securities account name-ISIN
                            key = f"{securities_account_name}-{investment}"
                            geneva_dict[key] = geneva_row['Traded Quantity']
                        
                        # Match and reconcile manually
                        reconciled_records = []
                        for _, holding_row in holding_df.iterrows():
                            securities_account_name = str(holding_row['Securities account name']).strip()
                            isin = str(holding_row['ISIN']).strip()
                            
                            # Create key: Securities account name-ISIN
                            key = f"{securities_account_name}-{isin}"
                            
                            # Check if this key exists in Geneva
                            if key in geneva_dict:
                                traded_quantity = geneva_dict[key]
                                traded_balance = holding_row['Traded balance']
                                
                                # Calculate difference: Traded Quantity - Traded balance
                                difference = traded_quantity - traded_balance
                                
                                # Create reconciled record
                                record = {
                                    'Value date as at': holding_row['Value date as at'],
                                    'Service location': holding_row['Service location'],
                                    'Securities account name': securities_account_name,
                                    'Securities account number': holding_row['Securities account number'],
                                    'Security type': holding_row['Security type'],
                                    'Security name': holding_row['Security name'],
                                    'ISIN': isin,
                                    'Place of settlement': holding_row['Place of settlement'],
                                    'Settled balance': holding_row['Settled balance'],
                                    'Anticipated debits': holding_row['Anticipated debits'],
                                    'Available balance': holding_row['Available balance'],
                                    'Anticipated credits': holding_row['Anticipated credits'],
                                    'Traded balance': traded_balance,
                                    'Security currency': holding_row['Security currency'],
                                    'Security price': holding_row['Security price'],
                                    'Indicative settled value': holding_row['Indicative settled value'],
                                    'Indicative traded value': holding_row['Indicative traded value'],
                                    'Indicative settled value in INR': holding_row['Indicative settled value in INR'],
                                    'Indicative traded value in INR': holding_row['Indicative traded value in INR'],
                                    'Traded Quantity': traded_quantity,
                                    'Difference': difference
                                }
                                
                                reconciled_records.append(record)
                        
                        # Add to Format 2 results
                        if reconciled_records:
                            format_2_results.extend(reconciled_records)
                    elif format_type == BHAVCOPY:
                        # Handle BhavCopy reconciliation with Geneva - Manual implementation
                        
                        # Create BhavCopy dict: ISIN -> BhavCopy data
                        bhavcopy_dict = {}
                        for _, row in holding_df.iterrows():
                            isin = row['ISIN']
                            bhavcopy_dict[isin] = {
                                'TradDt': row['TradDt'],
                                'BizDt': row['BizDt'],
                                'ISIN': row['ISIN'],
                                'TckrSymb': row['TckrSymb'],
                                'SctySrs': row['SctySrs'],
                                'OpnPric': row['OpnPric'],
                                'HghPric': row['HghPric'],
                                'LwPric': row['LwPric'],
                                'ClsPric': row['ClsPric'],
                                'LastPric': row['LastPric']
                            }
                        
                        # Match and reconcile manually
                        reconciled_records = []
                        for _, geneva_row in geneva_data.iterrows():
                            investment = geneva_row['Investment']
                            
                            # Check if this ISIN exists in BhavCopy
                            if investment in bhavcopy_dict:
                                bhavcopy_data = bhavcopy_dict[investment]
                                
                                # Calculate difference: Market Price Local - ClsPric
                                market_price = geneva_row['Market Price Local']
                                cls_pric = bhavcopy_data['ClsPric']
                                difference = market_price - cls_pric
                                
                                # Create reconciled record
                                record = {
                                    'Portfolio': geneva_row['Portfolio'],
                                    'Investment': geneva_row['Investment'],
                                    'Traded Quantity': geneva_row['Traded Quantity'],
                                    'Market Price Local': market_price,
                                    'TradDt': bhavcopy_data['TradDt'],
                                    'BizDt': bhavcopy_data['BizDt'],
                                    'ISIN': bhavcopy_data['ISIN'],
                                    'TckrSymb': bhavcopy_data['TckrSymb'],
                                    'SctySrs': bhavcopy_data['SctySrs'],
                                    'OpnPric': bhavcopy_data['OpnPric'],
                                    'HghPric': bhavcopy_data['HghPric'],
                                    'LwPric': bhavcopy_data['LwPric'],
                                    'ClsPric': cls_pric,
                                    'LastPric': bhavcopy_data['LastPric'],
                                    'Difference': difference
                                }
                                
                                reconciled_records.append(record)
                        
                        # Add to BhavCopy results
                        if reconciled_records:
                            bhavcopy_results.extend(reconciled_records)
                
                # Export to Excel - Separate sheets for each format
                with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
                    sheets_created = 0
                    
                    # Write Format 1 reconciliation - DBS Quantity Reconciliation
                    if format_1_results:
                        format_1_df = pd.DataFrame(format_1_results)
                        format_1_df.to_excel(writer, sheet_name='DBS_QTY_RECON', index=False)
                        self.format_excel_sheet(writer.sheets['DBS_QTY_RECON'], len(format_1_df), len(format_1_df.columns))
                        sheets_created += 1
                    
                    # Write Format 2 reconciliation - BNP Quantity Reconciliation
                    if format_2_results:
                        format_2_df = pd.DataFrame(format_2_results)
                        format_2_df.to_excel(writer, sheet_name='BNP_QTY_RECON', index=False)
                        self.format_excel_sheet(writer.sheets['BNP_QTY_RECON'], len(format_2_df), len(format_2_df.columns))
                        sheets_created += 1
                    
                    # Write BhavCopy reconciliation - Price Reconciliation
                    if bhavcopy_results:
                        bhavcopy_df = pd.DataFrame(bhavcopy_results)
                        bhavcopy_df.to_excel(writer, sheet_name='Price_Recon', index=False)
                        self.format_excel_sheet(writer.sheets['Price_Recon'], len(bhavcopy_df), len(bhavcopy_df.columns))
                        sheets_created += 1

                messagebox.showinfo("Success", f"Results exported to:\n{os.path.basename(export_path)}\n\nSheets created: {sheets_created}")

                self.status_var.set(f"Results exported to {os.path.basename(export_path)}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\n{str(e)}")
            self.status_var.set("Export failed")

# create a F&O Reconciliation seperate frame
# first filename
CDS_HOLDINGS_HEADER = ['Source.Name','Date','QuantityUnit','Exchange','ClientCode','TradingCode','UnderlyingCode','ClientName','UnderlyingName','InstrumentType','ExpiryDate','OptionType','StrikePrice','OpenBuy','OpenSell','TradedBuy','DayBuyValue','TradedSell','DaySellValue','ExcerciseQty','AllocationQty','NetBuy','NetSell','ContractSettlementPrice','Settlement Price','BloombergCodes','Concatenate','Fund Name In Geneva','UniqueCode','Netbuy-Netsell']


REGULAR_HOLDINGS_HEADERS = [
    'Source.Name', 'Date', 'Exchange', 'ClientName', 'ClientCode', 'OptionType',
    'UnderlyingCode', 'UnderlyingName', 'InstrumentType', 'StrikePrice', 'ExpiryDate',
    'OpenBuy', 'OpenSell', 'TradedSell', 'TradedBuy', 'DayBuyValue', 'DaySellValue',
    'ExcerciseQty', 'AllocationQty', 'NetBuy', 'NetSell', 'ContractSettlementPrice',
    'ClosingPrice', 'BloombergCD', 'Concatenate', 'Fund Name In Geneva', 'UniqueCode',
    'NerBuy-Netsell'
]

GENEVA_HOLDINGS_HEADERS = [
    'Portfolio', 'Investment', 'Investment Description', 'Traded Quantity', 'Settled Quantity',
    'Currency', 'Unit Cost', 'Cost Local', 'Cost Book', 'Market Price Local', 'Market Value Local',
    'Market Value Book', 'Unrealized G/L Book', 'Accrued Interest Local', 'Cust Account',
    'Strategy', 'UniqueCode', 'Fund Name IN Custody Holding', 'Nse Price', 'Diff'
]

NSE_F_AND_O_BHAVCOPY = [
    'INSTRUMENT', 'SYMBOL', 'EXPIRY_DT', 'STRIKE_PR', 'OPTION_TYP', 'OPEN',
    'HIGH', 'LOW', 'CLOSE', 'SETTLE_PR', 'CONTRACTS', 'VAL_INLAKH',
    'OPEN_INT', 'CHG_IN_OI', 'TIMESTAMP', 'ConcatenateCode'
]


