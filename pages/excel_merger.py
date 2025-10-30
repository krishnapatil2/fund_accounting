import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import threading
from datetime import datetime
import xlwings as xw
import time
from dataclasses import dataclass
from typing import Optional

from my_app.pages.loading import LoadingSpinner
from .helper import read_file


@dataclass
class FileEntry:
    """Represents a file entry for merging"""
    file_label: tk.Label
    sheet_entry: tk.Entry
    order_entry: tk.Entry
    remove_button: tk.Button
    replace_button: tk.Button
    file_path: str
    source_sheet_name: Optional[str] = None  # None for regular files, sheet name for extracted sheets
    
    @property
    def is_extracted_sheet(self) -> bool:
        """Check if this is an extracted sheet"""
        return self.source_sheet_name is not None
    
    @property
    def display_name(self) -> str:
        """Get display name for the file"""
        filename = os.path.basename(self.file_path)
        if self.is_extracted_sheet:
            return f"{self.source_sheet_name} (from {filename})"
        return filename


class ExcelMergerPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg="#ecf0f1")

        # Title
        title = tk.Label(self, text="📊 Excel Merger", font=("Arial", 20, "bold"), bg="#ecf0f1", fg="#2c3e50")
        title.pack(pady=10)

        # Main frame
        main_frame = tk.Frame(self, bg="#ecf0f1")
        main_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Header labels
        header_frame = tk.Frame(main_frame, bg="#ecf0f1")
        header_frame.pack(fill="x", pady=(0, 10))

        tk.Label(header_frame, text="File Name", font=("Arial", 11, "bold"), bg="#ecf0f1", fg="#2c3e50").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        tk.Label(header_frame, text="Sheet Name", font=("Arial", 11, "bold"), bg="#ecf0f1", fg="#2c3e50").grid(row=0, column=1, padx=10, pady=5, sticky="w")
        tk.Label(header_frame, text="Order", font=("Arial", 11, "bold"), bg="#ecf0f1", fg="#2c3e50").grid(row=0, column=2, padx=10, pady=5, sticky="w")

        # File entries container (scrollable)
        entries_container = tk.Frame(main_frame, bg="#ecf0f1")
        entries_container.pack(fill="both", expand=True)

        self.entries_canvas = tk.Canvas(entries_container, bg="#ecf0f1", highlightthickness=0)
        self.entries_scrollbar = ttk.Scrollbar(entries_container, orient="vertical", command=self.entries_canvas.yview)
        self.entries_canvas.configure(yscrollcommand=self.entries_scrollbar.set)

        self.entries_frame = tk.Frame(self.entries_canvas, bg="#ecf0f1")
        self.entries_canvas.create_window((0, 0), window=self.entries_frame, anchor="nw")

        self.entries_canvas.pack(side="left", fill="both", expand=True)
        self.entries_scrollbar.pack(side="right", fill="y")

        # Update scrollregion when size changes
        def _on_frame_configure(event):
            self.entries_canvas.configure(scrollregion=self.entries_canvas.bbox("all"))
        self.entries_frame.bind("<Configure>", _on_frame_configure)
        
        # Mousewheel scroll support
        def _on_mousewheel(event):
            try:
                delta = -1 * (event.delta // 120)
            except Exception:
                delta = 1
            self.entries_canvas.yview_scroll(delta, "units")
        self.entries_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Buttons
        button_frame = tk.Frame(main_frame, bg="#ecf0f1")
        button_frame.pack(fill="x", pady=10)

        tk.Button(button_frame, text="Add Files", command=self._browse_files, 
                 bg="#4CAF50", fg="white", relief="flat", padx=15, pady=8, font=("Arial", 11, "bold")).pack(side="left", padx=(0, 5))
        tk.Button(button_frame, text="Browse Multiple", command=self._browse_multiple, 
                 bg="#9b59b6", fg="white", relief="flat", padx=15, pady=8, font=("Arial", 11, "bold")).pack(side="left", padx=(0, 5))
        tk.Button(button_frame, text="Extract Sheets", command=self._extract_sheets, 
                 bg="#e67e22", fg="white", relief="flat", padx=15, pady=8, font=("Arial", 11, "bold")).pack(side="left", padx=(0, 5))
        tk.Button(button_frame, text="Clear All", command=self._clear_all, 
                 bg="#e74c3c", fg="white", relief="flat", padx=15, pady=8, font=("Arial", 11, "bold")).pack(side="left", padx=(0, 10))
        tk.Button(button_frame, text="Merge Files", command=self._merge_files, 
                 bg="#2196F3", fg="white", relief="flat", padx=15, pady=8, font=("Arial", 11, "bold")).pack(side="left")

        # Status
        self.status_var = tk.StringVar(value="Use 'Add Files', 'Browse Multiple', or 'Extract Sheets' to add files")
        tk.Label(self, textvariable=self.status_var, font=("Arial", 10), bg="#ecf0f1", fg="#7f8c8d").pack(fill="x", padx=20, pady=(0, 10))

        # Data storage
        self.file_entries = []  # List of FileEntry objects

    def _browse_files(self):
        """Browse and add Excel files from different folders"""
        files = filedialog.askopenfilenames(
            title="Select Files (XLS, XLSX, CSV formats supported)",
            filetypes=[
                ["All Excel & CSV", "*.xlsx *.xls *.csv"],
                ["Excel Files (XLSX)", "*.xlsx"],
                ["Excel Files (XLS)", "*.xls"], 
                ["CSV Files", "*.csv"],
                ["All Files", "*.*"]
            ]
        )
        if not files:
            return

        # Get current number of entries to continue numbering
        current_count = len(self.file_entries)
        
        for i, file_path in enumerate(files, start=current_count + 1):
            filename = os.path.basename(file_path)
            base_name = os.path.splitext(filename)[0]
            
            # File name label
            file_label = tk.Label(self.entries_frame, text=filename, anchor="w", width=40, bg="#ecf0f1", fg="#2c3e50")
            file_label.grid(row=i, column=0, padx=10, pady=3, sticky="w")

            # Sheet name entry
            sheet_entry = tk.Entry(self.entries_frame, width=25)
            sheet_entry.insert(0, base_name)
            sheet_entry.grid(row=i, column=1, padx=10, pady=3)

            # Order entry
            order_entry = tk.Entry(self.entries_frame, width=8)
            order_entry.insert(0, str(i))
            order_entry.grid(row=i, column=2, padx=10, pady=3)

            # Action buttons
            remove_btn = tk.Button(self.entries_frame, text="Remove", bg="#e74c3c", fg="white", relief="flat",
                                   padx=8, pady=2, command=lambda idx=len(self.file_entries): self._remove_entry(idx))
            replace_btn = tk.Button(self.entries_frame, text="Replace", bg="#3498db", fg="white", relief="flat",
                                    padx=8, pady=2, command=lambda idx=len(self.file_entries): self._replace_entry(idx))
            remove_btn.grid(row=i, column=3, padx=6, pady=2)
            replace_btn.grid(row=i, column=4, padx=0, pady=2)

            # Create FileEntry object
            file_entry = FileEntry(
                file_label=file_label,
                sheet_entry=sheet_entry,
                order_entry=order_entry,
                remove_button=remove_btn,
                replace_button=replace_btn,
                file_path=file_path,
                source_sheet_name=None  # Regular file, no specific sheet
            )
            self.file_entries.append(file_entry)

        self.status_var.set(f"Added {len(files)} more file(s). Total: {len(self.file_entries)} files")

    def _clear_all(self):
        """Clear all files from the list"""
        # Destroy all widgets
        for entry in self.file_entries:
            entry.file_label.destroy()
            entry.sheet_entry.destroy()
            entry.order_entry.destroy()
            entry.remove_button.destroy()
            entry.replace_button.destroy()
        self.file_entries.clear()
        
        self.status_var.set("Cleared all files")

    def _browse_multiple(self):
        """Browse and add files from multiple different folders in one session"""
        # Show dialog to select multiple files from different folders
        files = filedialog.askopenfilenames(
            title="Select Files from Multiple Folders (Hold Ctrl/Cmd to select from different folders)",
            filetypes=[
                ["All Excel & CSV", "*.xlsx *.xls *.csv"],
                ["Excel Files (XLSX)", "*.xlsx"],
                ["Excel Files (XLS)", "*.xls"], 
                ["CSV Files", "*.csv"],
                ["All Files", "*.*"]
            ]
        )
        if not files:
            return

        # Get current number of entries to continue numbering
        current_count = len(self.file_entries)
        
        # Group files by folder for better organization
        folder_groups = {}
        for file_path in files:
            folder = os.path.dirname(file_path)
            if folder not in folder_groups:
                folder_groups[folder] = []
            folder_groups[folder].append(file_path)
        
        # Add files maintaining folder grouping
        for folder, folder_files in folder_groups.items():
            for file_path in folder_files:
                current_count += 1
                filename = os.path.basename(file_path)
                base_name = os.path.splitext(filename)[0]
                
                # File name label (show folder info)
                folder_name = os.path.basename(folder)
                display_name = f"{filename} ({folder_name})" if len(folder_groups) > 1 else filename
                file_label = tk.Label(self.entries_frame, text=display_name, anchor="w", width=40, bg="#ecf0f1", fg="#2c3e50")
                file_label.grid(row=current_count, column=0, padx=10, pady=3, sticky="w")

                # Sheet name entry
                sheet_entry = tk.Entry(self.entries_frame, width=25)
                sheet_entry.insert(0, base_name)
                sheet_entry.grid(row=current_count, column=1, padx=10, pady=3)

                # Order entry
                order_entry = tk.Entry(self.entries_frame, width=8)
                order_entry.insert(0, str(current_count))
                order_entry.grid(row=current_count, column=2, padx=10, pady=3)

                # Action buttons
                remove_btn = tk.Button(self.entries_frame, text="Remove", bg="#e74c3c", fg="white", relief="flat",
                                       padx=8, pady=2, command=lambda idx=len(self.file_entries): self._remove_entry(idx))
                replace_btn = tk.Button(self.entries_frame, text="Replace", bg="#3498db", fg="white", relief="flat",
                                        padx=8, pady=2, command=lambda idx=len(self.file_entries): self._replace_entry(idx))
                remove_btn.grid(row=current_count, column=3, padx=6, pady=2)
                replace_btn.grid(row=current_count, column=4, padx=0, pady=2)

                # Create FileEntry object
                file_entry = FileEntry(
                    file_label=file_label,
                    sheet_entry=sheet_entry,
                    order_entry=order_entry,
                    remove_button=remove_btn,
                    replace_button=replace_btn,
                    file_path=file_path,
                    source_sheet_name=None  # Regular file, no specific sheet
                )
                self.file_entries.append(file_entry)

        folder_count = len(folder_groups)
        self.status_var.set(f"Added {len(files)} files from {folder_count} folder(s). Total: {len(self.file_entries)} files")

    def _extract_sheets(self):
        """Extract individual sheets from one Excel file"""
        # First, select the Excel file
        file_path = filedialog.askopenfilename(
            title="Select Excel File to Extract Sheets From (XLS, XLSX formats)",
            filetypes=[
                ["Excel Files (XLSX)", "*.xlsx"],
                ["Excel Files (XLS)", "*.xls"],
                ["All Excel Files", "*.xlsx *.xls"],
                ["All Files", "*.*"]
            ]
        )
        if not file_path:
            return

        try:
            # Get sheet names from the Excel file
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()

            if not sheet_names:
                messagebox.showwarning("No Sheets", "The selected Excel file has no sheets.")
                return

            # Show dialog to select which sheets to extract
            selected_sheets = self._show_sheet_selection_dialog(sheet_names, file_path)
            if not selected_sheets:
                return

            # Add selected sheets to the list
            self._add_extracted_sheets(file_path, selected_sheets)

        except Exception as e:
            messagebox.showerror("Error", f"Error reading Excel file: {str(e)}")

    def _show_sheet_selection_dialog(self, sheet_names, file_path):
        """Show dialog to select which sheets to extract"""
        dialog = tk.Toplevel(self)
        dialog.title("Select Sheets to Extract")
        dialog.geometry("500x400")
        dialog.configure(bg="#ecf0f1")
        dialog.transient(self)
        dialog.grab_set()

        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (400 // 2)
        dialog.geometry(f"500x400+{x}+{y}")

        # Title
        title = tk.Label(dialog, text=f"Select Sheets from: {os.path.basename(file_path)}", 
                       font=("Arial", 14, "bold"), bg="#ecf0f1", fg="#2c3e50")
        title.pack(pady=10)

        # Instructions
        instructions = tk.Label(dialog, text="Check the sheets you want to extract:", 
                               font=("Arial", 10), bg="#ecf0f1", fg="#7f8c8d")
        instructions.pack(pady=(0, 10))

        # Frame for checkboxes
        checkbox_frame = tk.Frame(dialog, bg="#ecf0f1")
        checkbox_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Scrollable frame for checkboxes
        canvas = tk.Canvas(checkbox_frame, bg="#ecf0f1")
        scrollbar = ttk.Scrollbar(checkbox_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#ecf0f1")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Variables to store checkbox states
        checkbox_vars = {}
        
        # Create checkboxes for each sheet
        for i, sheet_name in enumerate(sheet_names):
            var = tk.BooleanVar(value=True)  # Default to checked
            checkbox_vars[sheet_name] = var
            
            cb = tk.Checkbutton(scrollable_frame, text=sheet_name, variable=var, 
                               bg="#ecf0f1", fg="#2c3e50", font=("Arial", 10))
            cb.pack(anchor="w", pady=2)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Buttons
        button_frame = tk.Frame(dialog, bg="#ecf0f1")
        button_frame.pack(fill="x", padx=20, pady=10)

        result = []

        def select_all():
            for var in checkbox_vars.values():
                var.set(True)

        def select_none():
            for var in checkbox_vars.values():
                var.set(False)

        def ok_clicked():
            nonlocal result
            result = [sheet_name for sheet_name, var in checkbox_vars.items() if var.get()]
            dialog.destroy()

        tk.Button(button_frame, text="Select All", command=select_all, 
                 bg="#3498db", fg="white", relief="flat", padx=10, pady=5).pack(side="left", padx=(0, 5))
        tk.Button(button_frame, text="Select None", command=select_none, 
                 bg="#95a5a6", fg="white", relief="flat", padx=10, pady=5).pack(side="left", padx=(0, 10))
        tk.Button(button_frame, text="OK", command=ok_clicked, 
                 bg="#27ae60", fg="white", relief="flat", padx=20, pady=5, font=("Arial", 11, "bold")).pack(side="right", padx=(5, 0))
        tk.Button(button_frame, text="Cancel", command=dialog.destroy, 
                 bg="#e74c3c", fg="white", relief="flat", padx=20, pady=5, font=("Arial", 11, "bold")).pack(side="right")

        # Wait for dialog to close
        dialog.wait_window()
        return result

    def _add_extracted_sheets(self, file_path, selected_sheets):
        """Add extracted sheets to the file list"""
        current_count = len(self.file_entries)
        filename = os.path.basename(file_path)
        
        for i, sheet_name in enumerate(selected_sheets, start=current_count + 1):
            # File name label (show it's from extracted sheets)
            display_name = f"{sheet_name} (from {filename})"
            file_label = tk.Label(self.entries_frame, text=display_name, anchor="w", width=40, bg="#ecf0f1", fg="#2c3e50")
            file_label.grid(row=i, column=0, padx=10, pady=3, sticky="w")

            # Sheet name entry (use original sheet name)
            sheet_entry = tk.Entry(self.entries_frame, width=25)
            sheet_entry.insert(0, sheet_name)
            sheet_entry.grid(row=i, column=1, padx=10, pady=3)

            # Order entry
            order_entry = tk.Entry(self.entries_frame, width=8)
            order_entry.insert(0, str(i))
            order_entry.grid(row=i, column=2, padx=10, pady=3)

            # Action buttons
            remove_btn = tk.Button(self.entries_frame, text="Remove", bg="#e74c3c", fg="white", relief="flat",
                                   padx=8, pady=2, command=lambda idx=len(self.file_entries): self._remove_entry(idx))
            replace_btn = tk.Button(self.entries_frame, text="Replace", bg="#3498db", fg="white", relief="flat",
                                    padx=8, pady=2, command=lambda idx=len(self.file_entries): self._replace_entry(idx))
            remove_btn.grid(row=i, column=3, padx=6, pady=2)
            replace_btn.grid(row=i, column=4, padx=0, pady=2)

            # Create FileEntry object for extracted sheet
            file_entry = FileEntry(
                file_label=file_label,
                sheet_entry=sheet_entry,
                order_entry=order_entry,
                remove_button=remove_btn,
                replace_button=replace_btn,
                file_path=file_path,
                source_sheet_name=sheet_name  # Specific sheet name for extracted sheets
            )
            self.file_entries.append(file_entry)

        self.status_var.set(f"Extracted {len(selected_sheets)} sheet(s) from {filename}. Total: {len(self.file_entries)} items")

    def _validate_orders(self, orders):
        """Validate order numbers"""
        total = len(orders)
        unique_orders = set(orders)

        if len(unique_orders) != total:
            return "Duplicate order numbers detected."

        if not all(1 <= n <= total for n in orders):
            return f"Order numbers must be between 1 and {total}."

        if sorted(orders) != list(range(1, total + 1)):
            return f"Orders must be continuous (1 to {total})."

        return None  # valid

    def _merge_files(self):
        """Merge all selected files into a single Excel file"""
        if not self.file_entries:
            messagebox.showwarning("No Files", "Please add some files first.")
            return

        # Collect file data
        file_data = []
        for entry in self.file_entries:
            try:
                order = int(entry.order_entry.get())
            except ValueError:
                messagebox.showerror("Invalid Order", "Order must be a number.")
                return
            
            target_sheet_name = entry.sheet_entry.get().strip()
            if not target_sheet_name:
                messagebox.showerror("Missing Sheet Name", "Please enter all sheet names.")
                return
            
            file_data.append((order, target_sheet_name, entry.file_path, entry.source_sheet_name))

        # Validate order numbers
        order_list = [f[0] for f in file_data]
        validation_error = self._validate_orders(order_list)
        if validation_error:
            messagebox.showerror("Order Error", validation_error)
            return

        # Sort by order
        file_data.sort(key=lambda x: x[0])

        # Ask for output file
        output_path = filedialog.asksaveasfilename(
            title="Save Merged Excel File",
            defaultextension=".xlsx",
            filetypes=[["Excel Files", "*.xlsx"]],
            initialfile=f"Merged_Files_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not output_path:
            return

        # Show loading spinner
        loader = LoadingSpinner(self, text="Merging files...")

        def merge_task():
            try:
                self._perform_merge(output_path, file_data)
                loader.close()
                messagebox.showinfo("Success", f"Files merged successfully!\nSaved to: {output_path}")
                self.status_var.set("Merge completed successfully")
            except Exception as e:
                loader.close()
                raise
                messagebox.showerror("Merge Error", f"Error merging files: {str(e)}")
                self.status_var.set("Error during merge")

        # Run merge in background thread
        threading.Thread(target=merge_task, daemon=True).start()

    def _remove_entry(self, index):
        if index < 0 or index >= len(self.file_entries):
            return
        entry = self.file_entries.pop(index)
        try:
            entry.file_label.destroy()
            entry.sheet_entry.destroy()
            entry.order_entry.destroy()
            entry.remove_button.destroy()
            entry.replace_button.destroy()
        except Exception:
            pass
        self._reindex_entries()

    def _replace_entry(self, index):
        if index < 0 or index >= len(self.file_entries):
            return
        # Ask for a new file
        file_path = filedialog.askopenfilename(
            title="Select replacement file",
            filetypes=[
                ["All Excel & CSV", "*.xlsx *.xls *.csv"],
                ["Excel Files (XLSX)", "*.xlsx"],
                ["Excel Files (XLS)", "*.xls"],
                ["CSV Files", "*.csv"],
                ["All Files", "*.*"]
            ]
        )
        if not file_path:
            return
        filename = os.path.basename(file_path)
        base_name = os.path.splitext(filename)[0]
        entry = self.file_entries[index]
        entry.file_path = file_path
        entry.file_label.configure(text=filename)
        
        # Update sheet name entry with new filename
        entry.sheet_entry.delete(0, tk.END)
        entry.sheet_entry.insert(0, base_name)
        
        self.status_var.set(f"Replaced entry {index+1} with {filename}")

    def _reindex_entries(self):
        # Re-grid rows and reassign order numbers sequentially
        for idx, entry in enumerate(self.file_entries, start=1):
            try:
                entry.file_label.grid_configure(row=idx)
                entry.sheet_entry.grid_configure(row=idx)
                entry.order_entry.grid_configure(row=idx)
                entry.remove_button.grid_configure(row=idx)
                entry.replace_button.grid_configure(row=idx)
                # Update order entry text to reflect new index
                entry.order_entry.delete(0, tk.END)
                entry.order_entry.insert(0, str(idx))
            except Exception:
                pass

    def _perform_merge(self, output_path, file_data):
        """Perform the actual merge operation using xlwings for true formatting preservation"""
        # Create a new Excel workbook using xlwings
        app = xw.App(visible=False)
        app.display_alerts = False
        app.screen_updating = False
        
        try:
            # Create new workbook
            merged_wb = app.books.add()
            
            sheet_names_used = set()
            index_data = []
            first_sheet_replaced = False
            
            # Process each file in order
            for order, target_sheet_name, file_path, source_sheet_name in file_data:
                self._merge_single_file_xlwings(app, merged_wb, file_path, target_sheet_name, sheet_names_used, index_data, source_sheet_name, first_sheet_replaced)
                first_sheet_replaced = True
            
            # Create index sheet
            if index_data:
                self._create_index_sheet_xlwings(merged_wb, index_data)
            
            # Save the merged workbook
            merged_wb.save(output_path)
            merged_wb.close()
            
        finally:
            app.quit()

    def _merge_single_file_xlwings(self, app, merged_wb, file_path, target_sheet_name, sheet_names_used, index_data, source_sheet_name=None, first_sheet_replaced=False):
        """Merge a single file using xlwings copy_sheet_as_is approach"""
        try:
            # Open source workbook
            src_wb = app.books.open(file_path)
            filename = os.path.basename(file_path)
            file_ext = os.path.splitext(file_path)[1].lower()
            
            # Select source sheet
            if file_ext == ".csv":
                src_sheet = src_wb.sheets[0]
            else:
                # For Excel files, use specified sheet or first sheet
                if source_sheet_name:
                    # Extract specific sheet
                    src_sheet = src_wb.sheets[source_sheet_name]
                else:
                    # Use first sheet when source_sheet_name is None
                    # This happens when using "Add Files" button (not "Extract Sheets")
                    src_sheet = src_wb.sheets[0]
            
            # Generate unique sheet name
            unique_sheet_name = self._get_unique_sheet_name(target_sheet_name, sheet_names_used)
            sheet_names_used.add(unique_sheet_name)
            
            # Copy sheet as-is (with all formatting)
            if not first_sheet_replaced:
                # For the first file, we need to replace the default "Sheet1"
                default_sheet_name = merged_wb.sheets[0].name
                src_sheet.api.Copy(Before=merged_wb.sheets[0].api)
                # Wait for Excel to create the new sheet
                time.sleep(0.5)
                # The new sheet is now at position 0
                new_sheet = merged_wb.sheets[0]
                try:
                    new_sheet.name = unique_sheet_name
                    time.sleep(0.2)  # Wait for rename to complete
                except Exception:
                    # If name conflict, add suffix
                    new_sheet.name = f"{unique_sheet_name}_copy"
                    time.sleep(0.2)
                # Delete the original default sheet by name
                for sheet in merged_wb.sheets:
                    if sheet.name == default_sheet_name:
                        sheet.delete()
                        break
                time.sleep(0.2)
            else:
                # Add new sheet for subsequent files
                src_sheet.api.Copy(Before=merged_wb.sheets[0].api)
                # Wait for Excel to create the new sheet
                time.sleep(0.5)
                # Get the newly copied sheet (it's now at position 0)
                new_sheet = merged_wb.sheets[0]
                try:
                    new_sheet.name = unique_sheet_name
                    time.sleep(0.2)  # Wait for rename to complete
                except Exception:
                    # If name conflict, add suffix
                    new_sheet.name = f"{unique_sheet_name}_copy"
                    time.sleep(0.2)
            
            # Add to index data
            if source_sheet_name:
                description = f"Sheet '{source_sheet_name}' from {filename}"
            else:
                description = f"Data from {filename}"
            
            index_data.append({
                'sheet_name': unique_sheet_name,
                'source_file': filename,
                'description': description
            })
            
            # Close source workbook
            src_wb.close()
            
        except Exception as e:
            raise Exception(f"Error processing file {file_path}: {str(e)}")


    def _create_index_sheet_xlwings(self, merged_wb, index_data):
        """Create an index sheet using xlwings"""
        try:
            # Wait a bit to ensure all sheets are properly added
            time.sleep(0.5)
            
            # Create index sheet as the first sheet
            index_sheet = merged_wb.sheets.add("Index", before=merged_wb.sheets[0])
            
            # Wait for sheet to be created
            time.sleep(0.3)
            
            # Add header (row 2, column 2)
            header_cell = index_sheet.range(2, 2)
            header_cell.value = "Sheet Name"
            header_cell.font.bold = True
            header_cell.color = (68, 114, 196)  # Blue background
            header_cell.font.color = (255, 255, 255)  # White text
            
            # Add data rows (starting from row 3, column 2)
            for i, data in enumerate(index_data, start=3):
                # Sheet name with hyperlink
                sheet_cell = index_sheet.range(i, 2)
                sheet_cell.value = data['sheet_name']
                sheet_cell.add_hyperlink(f"#'{data['sheet_name']}'!A1", data['sheet_name'])
                sheet_cell.font.color = (0, 0, 255)  # Blue text for hyperlink
                sheet_cell.font.underline = True
            
            # Auto-fit columns
            index_sheet.autofit()
            
        except Exception as e:
            print(f"Warning: Could not create index sheet: {e}")

    def _get_unique_sheet_name(self, base_name, used_names):
        """Generate a unique sheet name"""
        # Use the provided name, but ensure uniqueness
        name = base_name[:31]  # Excel sheet name limit
        counter = 1
        original_name = name
        while name in used_names:
            name = f"{original_name}_{counter}"[:31]
            counter += 1
        return name

